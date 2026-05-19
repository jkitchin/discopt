"""Tests for the active-learning optimization stack.

Covers:
* Surrogate adapter: GP (return_std), BayesianRidge pipeline, plain
  LinearRegression bootstrap fallback, custom-protocol pass-through.
* Acquisition: EI sign for max/min, UCB vs LCB, steepest-ascent
  ignores σ.
* optimize_round: end-to-end loop on a 1D quadratic converges within
  6 rounds.
"""

from __future__ import annotations

from pathlib import Path

import numpy as np
import pytest
from discopt.doe.acquisition import (
    confidence_bound,
    expected_improvement,
    resolve_acquisition,
    steepest_ascent,
)
from discopt.doe.optimize import (
    OptimizationCriterion,
    optimize_round,
)
from discopt.doe.surrogate import (
    PRESETS,
    Surrogate,
    _SklearnUQAdapter,
    coerce_surrogate,
)
from discopt.doe.workbook import InputSpec, Workbook
from openpyxl import load_workbook as _lwb

# ──────────────────────────────────────────────────────────────────
# Surrogate adapter
# ──────────────────────────────────────────────────────────────────


def _quad_data(n=20, seed=0):
    rng = np.random.default_rng(seed)
    X = rng.uniform(-3.0, 3.0, size=(n, 1))
    y = (X**2).ravel() + 0.05 * rng.normal(size=n)
    return X, y


def test_preset_gp_uses_return_std():
    s = coerce_surrogate("gp")
    X, y = _quad_data()
    s.fit(X, y)
    assert s.mode == "return_std"
    mean, std = s.predict(np.array([[0.0]]))
    assert mean.shape == (1,)
    assert std.shape == (1,)
    assert std[0] >= 0.0


def test_preset_response_surface_uses_return_std():
    s = coerce_surrogate("response-surface")
    X, y = _quad_data()
    s.fit(X, y)
    assert s.mode == "return_std"
    mean, std = s.predict(np.array([[0.0], [2.0]]))
    assert np.all(std >= 0.0)
    # degree-2 polynomial should fit y = x^2 well
    assert mean[0] < 0.5  # near 0 at x=0
    assert abs(mean[1] - 4.0) < 0.5  # near 4 at x=2


def test_adapter_wraps_arbitrary_sklearn_estimator():
    from sklearn.gaussian_process import GaussianProcessRegressor

    gp = GaussianProcessRegressor(normalize_y=True)
    s = coerce_surrogate(gp)
    assert isinstance(s, _SklearnUQAdapter)
    X, y = _quad_data()
    s.fit(X, y)
    assert s.mode == "return_std"


def test_adapter_bootstrap_fallback_for_linear_regression():
    from sklearn.linear_model import LinearRegression

    s = coerce_surrogate(LinearRegression())
    X, y = _quad_data()
    s.fit(X, y)
    assert s.mode == "bootstrap"
    mean, std = s.predict(np.array([[0.0], [2.0]]))
    assert np.all(std > 0.0)  # bootstrap σ should be positive


def test_coerce_unknown_preset_raises():
    with pytest.raises(ValueError, match="unknown surrogate preset"):
        coerce_surrogate("not-a-real-preset")


def test_coerce_garbage_raises():
    with pytest.raises(TypeError):
        coerce_surrogate(42)


def test_custom_protocol_passes_through():
    class Custom:
        _is_discopt_surrogate = True

        def fit(self, X, y):
            self._mean = float(np.mean(y))
            return self

        def predict(self, X):
            n = len(X)
            return np.full(n, self._mean), np.ones(n)

    s = coerce_surrogate(Custom())
    assert isinstance(s, Surrogate)
    X, y = _quad_data()
    s.fit(X, y)
    mean, std = s.predict(np.array([[0.0]]))
    assert std[0] == 1.0  # would be 0 if the adapter wrapped this


def test_presets_listed():
    assert set(PRESETS) >= {"gp", "response-surface"}


# ──────────────────────────────────────────────────────────────────
# Acquisition functions
# ──────────────────────────────────────────────────────────────────


def _fit_gp_quadratic():
    s = coerce_surrogate("gp")
    X, y = _quad_data()
    s.fit(X, y)
    return s, X, y


def test_ei_argmax_at_minimum_for_minimization():
    s, X, y = _fit_gp_quadratic()
    Xc = np.linspace(-3, 3, 61).reshape(-1, 1)
    ei = expected_improvement(s, Xc, y_best=float(y.min()), direction=-1)
    assert abs(Xc[ei.argmax(), 0]) < 0.5  # close to x=0


def test_ei_argmax_at_boundary_for_maximization():
    s, X, y = _fit_gp_quadratic()
    Xc = np.linspace(-3, 3, 61).reshape(-1, 1)
    ei = expected_improvement(s, Xc, y_best=float(y.max()), direction=1)
    assert abs(Xc[ei.argmax(), 0]) > 2.5  # near a boundary


def test_ei_nonnegative():
    s, X, y = _fit_gp_quadratic()
    Xc = np.linspace(-3, 3, 31).reshape(-1, 1)
    for d in (1, -1):
        ei = expected_improvement(s, Xc, y_best=float(y.mean()), direction=d)
        assert np.all(ei >= -1e-9)


def test_ucb_lcb_direction():
    s, X, y = _fit_gp_quadratic()
    Xc = np.linspace(-3, 3, 31).reshape(-1, 1)
    ucb = confidence_bound(s, Xc, direction=1, kappa=2.0)
    lcb = confidence_bound(s, Xc, direction=-1, kappa=2.0)
    # UCB favors high μ; LCB favors low μ. Their argmaxes differ.
    assert ucb.argmax() != lcb.argmax()


def test_steepest_ascent_ignores_sigma():
    s, X, y = _fit_gp_quadratic()
    Xc = np.linspace(-3, 3, 31).reshape(-1, 1)
    sa_max = steepest_ascent(s, Xc, direction=1)
    sa_min = steepest_ascent(s, Xc, direction=-1)
    assert sa_max.argmax() != sa_min.argmax()
    # Minimization should pick near x=0 (where μ is smallest)
    assert abs(Xc[sa_min.argmax(), 0]) < 0.5


def test_resolve_acquisition_strings():
    assert resolve_acquisition("expected_improvement") is expected_improvement
    assert resolve_acquisition("ei") is expected_improvement
    assert resolve_acquisition("ucb") is confidence_bound
    assert resolve_acquisition("steepest_ascent") is steepest_ascent


def test_resolve_acquisition_callable_passthrough():
    def my_acq(s, X, **kw):
        return np.zeros(len(X))

    assert resolve_acquisition(my_acq) is my_acq


def test_resolve_acquisition_unknown_raises():
    with pytest.raises(ValueError, match="unknown acquisition"):
        resolve_acquisition("nope")


# ──────────────────────────────────────────────────────────────────
# optimize_round
# ──────────────────────────────────────────────────────────────────


def _make_seeded_workbook(tmp_path: Path, init_xs, truth_fn):
    path = tmp_path / "opt.xlsx"
    Workbook.create(
        path,
        template=None,
        template_args={},
        input_specs=[InputSpec("x", -5.0, 5.0)],
        criterion="custom",
        measurement_error=0.05,
        seed=0,
        response_name="y",
    )
    wb = Workbook.open(path)
    wb.append_runs(0, [{"x": float(x)} for x in init_xs])
    wb.save()
    book = _lwb(path)
    sh = book["runs"]
    for i, x in enumerate(init_xs, start=2):
        sh.cell(row=i, column=4, value=float(truth_fn(x)))
    book.save(path)
    return path


def _fill_responses(path, run_ids, designs, truth_fn):
    book = _lwb(path)
    sh = book["runs"]
    by_id = {rid: d for rid, d in zip(run_ids, designs)}
    for row in sh.iter_rows(min_row=2):
        rid = row[0].value
        if rid in by_id:
            row[3].value = float(truth_fn(by_id[rid]["x"]))
    book.save(path)


def test_optimize_round_converges_to_max(tmp_path):
    # y = -(x-2)^2 + 3, max at x=2, y=3
    rng = np.random.default_rng(0)

    def truth(x):
        return -((x - 2.0) ** 2) + 3.0 + 0.05 * rng.normal()

    init_xs = np.array([-4.0, -1.0, 1.0, 4.0])
    path = _make_seeded_workbook(tmp_path, init_xs, truth)

    for rnd in range(6):
        result = optimize_round(
            workbook=path,
            criterion=OptimizationCriterion.MAXIMIZE,
            surrogate="gp",
            acquisition="expected_improvement",
            batch_size=2,
            seed=rnd,
        )
        _fill_responses(path, result.new_run_ids, result.next_designs, truth)

    final = Workbook.open(path).completed_runs()
    ys = [float(r["y"]) for r in final]
    best_x = float(final[max(range(len(ys)), key=lambda i: ys[i])]["x"])
    assert abs(best_x - 2.0) < 0.5
    assert max(ys) > 2.5


def test_optimize_round_converges_to_min(tmp_path):
    rng = np.random.default_rng(1)

    def truth(x):
        return (x - 1.0) ** 2 + 0.05 * rng.normal()

    init_xs = np.array([-4.0, -2.0, 2.0, 4.0])
    path = _make_seeded_workbook(tmp_path, init_xs, truth)
    for rnd in range(6):
        result = optimize_round(
            workbook=path,
            criterion=OptimizationCriterion.MINIMIZE,
            surrogate="gp",
            acquisition="expected_improvement",
            batch_size=2,
            seed=rnd,
        )
        _fill_responses(path, result.new_run_ids, result.next_designs, truth)

    final = Workbook.open(path).completed_runs()
    ys = [float(r["y"]) for r in final]
    best_x = float(final[min(range(len(ys)), key=lambda i: ys[i])]["x"])
    assert abs(best_x - 1.0) < 0.5
    assert min(ys) < 0.5


def test_optimize_round_diverse_batch(tmp_path):
    """Two-point batch should not return duplicate designs."""
    rng = np.random.default_rng(2)

    def truth(x):
        return -((x - 0.5) ** 2) + 0.05 * rng.normal()

    init_xs = np.array([-3.0, 0.0, 3.0])
    path = _make_seeded_workbook(tmp_path, init_xs, truth)
    result = optimize_round(
        workbook=path,
        criterion=OptimizationCriterion.MAXIMIZE,
        surrogate="gp",
        acquisition="expected_improvement",
        batch_size=4,
        seed=0,
    )
    xs = [d["x"] for d in result.next_designs]
    assert len(set(round(x, 6) for x in xs)) == len(xs)


def test_optimize_round_empty_workbook_raises(tmp_path):
    path = tmp_path / "empty.xlsx"
    Workbook.create(
        path,
        template=None,
        template_args={},
        input_specs=[InputSpec("x", -1.0, 1.0)],
        criterion="custom",
        measurement_error=1.0,
        seed=0,
        response_name="y",
    )
    with pytest.raises(ValueError, match="no completed runs"):
        optimize_round(workbook=path, surrogate="gp")


def test_optimize_round_user_supplied_estimator(tmp_path):
    """Custom sklearn estimator is auto-wrapped."""
    from sklearn.gaussian_process import GaussianProcessRegressor
    from sklearn.gaussian_process.kernels import RBF

    rng = np.random.default_rng(3)

    def truth(x):
        return -((x - 1.0) ** 2) + 0.05 * rng.normal()

    init_xs = np.array([-3.0, 0.0, 3.0, 1.5])
    path = _make_seeded_workbook(tmp_path, init_xs, truth)
    custom_gp = GaussianProcessRegressor(kernel=RBF(length_scale=1.0), normalize_y=True)
    result = optimize_round(
        workbook=path,
        criterion=OptimizationCriterion.MAXIMIZE,
        surrogate=custom_gp,
        acquisition="ucb",
        acquisition_kwargs={"kappa": 1.5},
        batch_size=1,
        seed=0,
    )
    assert len(result.next_designs) == 1
    assert result.surrogate_mode == "return_std"
