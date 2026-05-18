"""Excel-workbook state layer for the ``discopt doe`` CLI.

A single ``.xlsx`` file is the source of truth for a DoE campaign.
The workbook is portable between the lab bench (where experimentalists
fill in measured responses) and the CLI (which estimates parameters
and proposes the next batch). All CLI verbs after ``new`` need only
the workbook path — template metadata, prior FIM, and per-run state
all live inside the file.

Sheets
------
``runs``
    Columns: ``run_id``, ``batch``, ``<input_1>``, …, ``<response>``,
    ``measured_at``. One row per recommended experiment. The CLI
    treats a row as "completed" when its response column is non-empty.

``metadata``
    Key/value sheet. Includes ``discopt_version``, ``template``,
    ``template_args`` (JSON), ``input_specs`` (JSON list of
    ``{name, lb, ub}``), ``criterion``, ``measurement_error``, ``seed``,
    ``response_name``, ``module_callable`` (escape-hatch only), and
    ``param_initial_guess`` (JSON, escape-hatch only). Sufficient to
    re-build the :class:`~discopt.estimate.Experiment` in
    :meth:`Workbook.rebuild_experiment`.

``parameters``
    Written by ``fit`` and ``extend``. Columns: ``name``, ``estimate``,
    ``std_error``, ``ci_lower_95``, ``ci_upper_95``, ``updated_at``.

``fim``
    The latest cumulative Fisher Information Matrix, labelled
    rows/columns. Used by ``extend`` as the ``prior_fim`` for the next
    optimal-design call.

``history``
    Append-only log of CLI commands run against this workbook.

``instructions``
    Human-readable workflow guide, one styled block per row in column A
    (title, section headers, body, code). Rendered by
    :func:`_write_instructions_sheet`.

``anova``
    Regression diagnostics written by ``fit``: per-parameter coefficient
    table (estimate, std error, t-statistic, p-value, 95% CI), overall
    ANOVA decomposition (regression / residual / total sums of squares,
    F-statistic, p-value), and fit-summary statistics (R², adjusted R²,
    RMSE). Only populated for built-in templates (the OLS path).
"""

from __future__ import annotations

import datetime as _dt
import json
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import numpy as np

from discopt.estimate import Experiment

# Sheet names — single source of truth.
SHEET_RUNS = "runs"
SHEET_METADATA = "metadata"
SHEET_PARAMETERS = "parameters"
SHEET_FIM = "fim"
SHEET_HISTORY = "history"
SHEET_INSTRUCTIONS = "instructions"
SHEET_ANOVA = "anova"

# Instructions sheet layout: a list of (style, text) blocks rendered top-to-bottom.
# Styles: "title" (large bold), "h2" (bold section header), "body" (wrapped prose),
# "code" (monospace, no wrap). Each block becomes one cell in column A.
_INSTRUCTIONS_BLOCKS: list[tuple[str, str]] = [
    ("title", "discopt DoE campaign workbook"),
    ("h2", "Workflow"),
    (
        "body",
        "1. The 'runs' sheet lists the experiments to perform. For each row, "
        "run the experiment at the listed input conditions and fill in the "
        "response column (and optionally 'measured_at'). Leave rows you "
        "haven't done yet blank.",
    ),
    ("body", "2. Save the workbook."),
    ("body", "3. From the command line, check status:"),
    ("code", "    discopt doe status <this-file.xlsx>"),
    ("body", "   then estimate parameters from the completed rows:"),
    ("code", "    discopt doe fit <this-file.xlsx>"),
    ("body", "   Results are written to the 'parameters' and 'fim' sheets."),
    ("body", "4. To request more recommended runs:"),
    ("code", "    discopt doe extend <this-file.xlsx> --n M"),
    ("body", "   New rows are appended to 'runs' with the next batch number."),
    ("h2", "Hands off"),
    (
        "body",
        "Do not edit the 'metadata', 'fim', or 'history' sheets by hand. "
        "The CLI rewrites them and any manual edits will be lost.",
    ),
    ("h2", "Sheets in this workbook"),
    ("body", "runs — experiments to perform; you fill in the response column."),
    ("body", "metadata — campaign configuration (template, inputs, criterion, seed, ...)."),
    ("body", "parameters — fitted parameter estimates, std errors, and 95% CIs."),
    (
        "body",
        "anova — regression report: coefficient table (estimate, SE, t, p, 95% CI), "
        "ANOVA decomposition, R², adjusted R², RMSE.",
    ),
    ("body", "fim — cumulative Fisher Information Matrix (reused by 'extend')."),
    ("body", "history — append-only log of CLI commands run against this workbook."),
]


def _write_instructions_sheet(sheet: Any) -> None:
    """Render the instructions blocks into ``sheet`` with readable formatting."""
    from openpyxl.styles import Alignment, Font

    # Single wide column, generous height so wrapped prose isn't clipped.
    sheet.column_dimensions["A"].width = 100

    title_font = Font(name="Calibri", size=16, bold=True)
    h2_font = Font(name="Calibri", size=12, bold=True)
    body_font = Font(name="Calibri", size=11)
    code_font = Font(name="Consolas", size=10)
    wrap = Alignment(wrap_text=True, vertical="top")
    nowrap = Alignment(wrap_text=False, vertical="top")

    # Heuristic row-height per style; openpyxl can't auto-fit wrapped text.
    row_height = {"title": 28, "h2": 22, "body": None, "code": 18}
    font_for = {"title": title_font, "h2": h2_font, "body": body_font, "code": code_font}
    align_for = {"title": nowrap, "h2": nowrap, "body": wrap, "code": nowrap}

    # Width in characters used for body row-height estimation.
    col_chars = 95

    for idx, (style, text) in enumerate(_INSTRUCTIONS_BLOCKS, start=1):
        cell = sheet.cell(row=idx, column=1, value=text)
        cell.font = font_for[style]
        cell.alignment = align_for[style]
        if style == "body":
            # Estimate wrapped line count; ~16pt per line.
            lines = max(1, (len(text) + col_chars - 1) // col_chars)
            sheet.row_dimensions[idx].height = 16 * lines + 4
        else:
            sheet.row_dimensions[idx].height = row_height[style]


@dataclass
class InputSpec:
    name: str
    lb: float
    ub: float

    def to_dict(self) -> dict[str, Any]:
        return {"name": self.name, "lb": self.lb, "ub": self.ub}

    @classmethod
    def from_dict(cls, d: dict[str, Any]) -> "InputSpec":
        return cls(name=d["name"], lb=float(d["lb"]), ub=float(d["ub"]))


def _require_openpyxl():
    try:
        import openpyxl  # noqa: F401
    except ImportError as e:
        raise ImportError(
            "discopt doe needs openpyxl. Install with: pip install 'discopt[doe]'"
        ) from e


def _now_iso() -> str:
    return _dt.datetime.now(_dt.timezone.utc).replace(microsecond=0).isoformat()


def _coerce_float(v: Any) -> float | None:
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


class Workbook:
    """Read/write wrapper for a discopt DoE campaign workbook."""

    def __init__(self, path: Path, wb: Any) -> None:
        self.path = Path(path)
        self._wb = wb

    # ------------------------------------------------------------------
    # Construction / open / save
    # ------------------------------------------------------------------

    @classmethod
    def create(
        cls,
        path: Path | str,
        *,
        template: str | None,
        template_args: dict[str, Any],
        input_specs: list[InputSpec],
        criterion: str,
        measurement_error: float,
        seed: int,
        response_name: str,
        module_callable: str | None = None,
        param_initial_guess: dict[str, float] | None = None,
    ) -> "Workbook":
        _require_openpyxl()
        from openpyxl import Workbook as _OpenpyxlWorkbook

        path = Path(path)
        wb = _OpenpyxlWorkbook()
        # Remove default sheet
        default = wb.active
        wb.remove(default)

        # Create all sheets in a stable order
        wb.create_sheet(SHEET_INSTRUCTIONS)
        wb.create_sheet(SHEET_RUNS)
        wb.create_sheet(SHEET_METADATA)
        wb.create_sheet(SHEET_PARAMETERS)
        wb.create_sheet(SHEET_ANOVA)
        wb.create_sheet(SHEET_FIM)
        wb.create_sheet(SHEET_HISTORY)

        # Instructions
        _write_instructions_sheet(wb[SHEET_INSTRUCTIONS])

        # Metadata
        import discopt as _discopt

        meta_rows: list[tuple[str, Any]] = [
            ("discopt_version", _discopt.__version__),
            ("created_at", _now_iso()),
            ("template", template or ""),
            ("template_args", json.dumps(template_args, sort_keys=True)),
            ("input_specs", json.dumps([s.to_dict() for s in input_specs], sort_keys=True)),
            ("criterion", criterion),
            ("measurement_error", float(measurement_error)),
            ("seed", int(seed)),
            ("response_name", response_name),
            ("module_callable", module_callable or ""),
            (
                "param_initial_guess",
                json.dumps(param_initial_guess or {}, sort_keys=True),
            ),
        ]
        meta_sheet = wb[SHEET_METADATA]
        meta_sheet.append(["key", "value"])
        for k, v in meta_rows:
            meta_sheet.append([k, v])

        # Runs header row
        runs_sheet = wb[SHEET_RUNS]
        header = (
            ["run_id", "batch"] + [s.name for s in input_specs] + [response_name, "measured_at"]
        )
        runs_sheet.append(header)

        # Parameters header
        params_sheet = wb[SHEET_PARAMETERS]
        params_sheet.append(
            ["name", "estimate", "std_error", "ci_lower_95", "ci_upper_95", "updated_at"]
        )

        # History header
        hist_sheet = wb[SHEET_HISTORY]
        hist_sheet.append(["timestamp", "command", "args", "note"])

        # FIM sheet starts empty; populated by fit/extend.

        path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(path)
        return cls(path, wb)

    @classmethod
    def open(cls, path: Path | str) -> "Workbook":
        _require_openpyxl()
        from openpyxl import load_workbook

        path = Path(path)
        if not path.exists():
            raise FileNotFoundError(f"workbook not found: {path}")
        wb = load_workbook(path)
        required = {
            SHEET_RUNS,
            SHEET_METADATA,
            SHEET_PARAMETERS,
            SHEET_FIM,
            SHEET_HISTORY,
        }
        missing = required - set(wb.sheetnames)
        if missing:
            raise ValueError(f"workbook {path} is missing required sheets: {sorted(missing)}")
        return cls(path, wb)

    def save(self) -> None:
        self._wb.save(self.path)

    # ------------------------------------------------------------------
    # Metadata
    # ------------------------------------------------------------------

    def metadata(self) -> dict[str, Any]:
        sheet = self._wb[SHEET_METADATA]
        out: dict[str, Any] = {}
        # Skip header row.
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            out[str(row[0])] = row[1]
        return out

    def input_specs(self) -> list[InputSpec]:
        raw = self.metadata().get("input_specs") or "[]"
        return [InputSpec.from_dict(d) for d in json.loads(raw)]

    def response_name(self) -> str:
        return str(self.metadata().get("response_name") or "y")

    def template_name(self) -> str | None:
        v = self.metadata().get("template")
        return str(v) if v else None

    def template_args(self) -> dict[str, Any]:
        raw = self.metadata().get("template_args") or "{}"
        result: dict[str, Any] = json.loads(raw)
        return result

    def module_callable(self) -> str | None:
        v = self.metadata().get("module_callable")
        return str(v) if v else None

    def param_initial_guess(self) -> dict[str, float]:
        raw = self.metadata().get("param_initial_guess") or "{}"
        out: dict[str, float] = {}
        for k, v in json.loads(raw).items():
            out[str(k)] = float(v)
        return out

    def criterion(self) -> str:
        return str(self.metadata().get("criterion") or "determinant")

    def measurement_error(self) -> float:
        return float(self.metadata().get("measurement_error") or 1.0)

    def seed(self) -> int:
        return int(self.metadata().get("seed") or 42)

    # ------------------------------------------------------------------
    # Runs
    # ------------------------------------------------------------------

    def _runs_headers(self) -> list[str]:
        sheet = self._wb[SHEET_RUNS]
        first = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        return [str(c) for c in (first or [])]

    def _input_column_names(self) -> list[str]:
        return [s.name for s in self.input_specs()]

    def append_runs(self, batch_idx: int, runs: list[dict[str, float]]) -> list[int]:
        """Append a batch of pending runs to the workbook. Returns the new run_ids."""
        sheet = self._wb[SHEET_RUNS]
        input_names = self._input_column_names()
        # Determine next run_id
        existing_ids = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row and row[0] is not None:
                try:
                    existing_ids.append(int(row[0]))
                except (TypeError, ValueError):
                    continue
        next_id = (max(existing_ids) + 1) if existing_ids else 1
        new_ids: list[int] = []
        for run in runs:
            row = [next_id, int(batch_idx)]
            for nm in input_names:
                row.append(float(run[nm]))
            row.append(None)  # response (blank => pending)
            row.append(None)  # measured_at
            sheet.append(row)
            new_ids.append(next_id)
            next_id += 1
        return new_ids

    def all_runs(self) -> list[dict[str, Any]]:
        """Return every run as a dict (including pending rows)."""
        sheet = self._wb[SHEET_RUNS]
        headers = self._runs_headers()
        out: list[dict[str, Any]] = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            out.append(dict(zip(headers, row)))
        return out

    def pending_runs(self) -> list[dict[str, Any]]:
        response = self.response_name()
        return [r for r in self.all_runs() if _coerce_float(r.get(response)) is None]

    def completed_runs(self) -> list[dict[str, Any]]:
        response = self.response_name()
        return [r for r in self.all_runs() if _coerce_float(r.get(response)) is not None]

    def next_batch_index(self) -> int:
        batches = [r.get("batch") for r in self.all_runs()]
        ints = [int(b) for b in batches if b is not None]
        return (max(ints) + 1) if ints else 1

    # ------------------------------------------------------------------
    # Parameters / FIM
    # ------------------------------------------------------------------

    def write_parameters(
        self,
        names: list[str],
        estimates: dict[str, float],
        std_errors: dict[str, float],
        confidence_intervals: dict[str, tuple[float, float]],
    ) -> None:
        sheet = self._wb[SHEET_PARAMETERS]
        # Clear existing data rows, keep header.
        if sheet.max_row > 1:
            sheet.delete_rows(2, sheet.max_row - 1)
        ts = _now_iso()
        for name in names:
            est = float(estimates.get(name, float("nan")))
            se = float(std_errors.get(name, float("nan")))
            lo, hi = confidence_intervals.get(name, (float("nan"), float("nan")))
            sheet.append([name, est, se, float(lo), float(hi), ts])

    def write_anova(
        self,
        *,
        coefficients: list[dict[str, Any]],
        anova_rows: list[dict[str, Any]],
        fit_summary: list[tuple[str, Any]],
    ) -> None:
        """Write the regression report (coefficients + ANOVA + summary) sheet.

        ``coefficients`` rows: ``name``, ``estimate``, ``std_error``,
        ``t_statistic``, ``p_value``, ``ci_lower_95``, ``ci_upper_95``.
        ``anova_rows`` rows: ``source``, ``ss``, ``df``, ``ms``,
        ``f_statistic``, ``p_value`` (last two may be ``None`` for the
        Residual / Total rows). ``fit_summary`` is an ordered list of
        ``(label, value)`` pairs.
        """
        from openpyxl.styles import Alignment, Font

        if SHEET_ANOVA not in self._wb.sheetnames:
            self._wb.create_sheet(SHEET_ANOVA)
        sheet = self._wb[SHEET_ANOVA]
        if sheet.max_row >= 1:
            sheet.delete_rows(1, sheet.max_row)

        # Column widths tuned for the longest expected content.
        widths = {"A": 22, "B": 16, "C": 8, "D": 16, "E": 14, "F": 14, "G": 14}
        for col, w in widths.items():
            sheet.column_dimensions[col].width = w

        h1 = Font(name="Calibri", size=14, bold=True)
        h2_font = Font(name="Calibri", size=11, bold=True)
        body = Font(name="Calibri", size=11)
        right = Alignment(horizontal="right")

        def _section(row: int, title: str) -> int:
            cell = sheet.cell(row=row, column=1, value=title)
            cell.font = h1
            return row + 1

        def _header(row: int, cols: list[str]) -> int:
            for j, name in enumerate(cols, start=1):
                c = sheet.cell(row=row, column=j, value=name)
                c.font = h2_font
            return row + 1

        def _num(row: int, col: int, value: Any, fmt: str = "0.0000") -> None:
            c = sheet.cell(row=row, column=col, value=value)
            c.font = body
            c.alignment = right
            if isinstance(value, (int, float)) and value is not None:
                c.number_format = fmt

        r = 1
        # --- Coefficients ---
        r = _section(r, "Coefficients")
        r = _header(
            r,
            [
                "name",
                "estimate",
                "std_error",
                "t_statistic",
                "p_value",
                "ci_lower_95",
                "ci_upper_95",
            ],
        )
        for coef in coefficients:
            sheet.cell(row=r, column=1, value=coef["name"]).font = body
            _num(r, 2, coef.get("estimate"), "0.000000")
            _num(r, 3, coef.get("std_error"), "0.0000E+00")
            _num(r, 4, coef.get("t_statistic"), "0.0000")
            _num(r, 5, coef.get("p_value"), "0.0000E+00")
            _num(r, 6, coef.get("ci_lower_95"), "0.000000")
            _num(r, 7, coef.get("ci_upper_95"), "0.000000")
            r += 1
        r += 1  # blank spacer

        # --- ANOVA table ---
        r = _section(r, "ANOVA")
        r = _header(r, ["source", "SS", "df", "MS", "F", "p_value"])
        for row in anova_rows:
            sheet.cell(row=r, column=1, value=row["source"]).font = body
            _num(r, 2, row.get("ss"), "0.000000")
            _num(r, 3, row.get("df"), "0")
            _num(r, 4, row.get("ms"), "0.000000")
            _num(r, 5, row.get("f_statistic"), "0.0000")
            _num(r, 6, row.get("p_value"), "0.0000E+00")
            r += 1
        r += 1

        # --- Fit summary ---
        r = _section(r, "Fit summary")
        for label, value in fit_summary:
            sheet.cell(row=r, column=1, value=label).font = body
            if isinstance(value, float):
                _num(r, 2, value, "0.000000")
            else:
                c = sheet.cell(row=r, column=2, value=value)
                c.font = body
                c.alignment = right
            r += 1

    def read_parameters(self) -> list[dict[str, Any]]:
        sheet = self._wb[SHEET_PARAMETERS]
        out: list[dict[str, Any]] = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            out.append(
                {
                    "name": str(row[0]),
                    "estimate": _coerce_float(row[1]),
                    "std_error": _coerce_float(row[2]),
                    "ci_lower_95": _coerce_float(row[3]),
                    "ci_upper_95": _coerce_float(row[4]),
                    "updated_at": row[5],
                }
            )
        return out

    def write_fim(self, fim: np.ndarray, names: list[str]) -> None:
        sheet = self._wb[SHEET_FIM]
        if sheet.max_row >= 1:
            sheet.delete_rows(1, sheet.max_row)
        sheet.append([""] + list(names))
        for i, name in enumerate(names):
            sheet.append([name] + [float(fim[i, j]) for j in range(len(names))])

    def read_fim(self) -> tuple[np.ndarray, list[str]] | None:
        sheet = self._wb[SHEET_FIM]
        rows = list(sheet.iter_rows(values_only=True))
        if not rows or len(rows) < 2:
            return None
        header = rows[0]
        names = [str(c) for c in header[1:] if c is not None]
        if not names:
            return None
        n = len(names)
        m = np.zeros((n, n), dtype=np.float64)
        for i, row in enumerate(rows[1 : 1 + n]):
            for j in range(n):
                val = _coerce_float(row[1 + j])
                m[i, j] = 0.0 if val is None else val
        return m, names

    # ------------------------------------------------------------------
    # History
    # ------------------------------------------------------------------

    def log(self, command: str, args: dict[str, Any] | None = None, note: str = "") -> None:
        sheet = self._wb[SHEET_HISTORY]
        sheet.append([_now_iso(), command, json.dumps(args or {}, sort_keys=True), note])

    # ------------------------------------------------------------------
    # Rebuild Experiment from metadata
    # ------------------------------------------------------------------

    def rebuild_experiment(self) -> tuple[Experiment, list[str]]:
        """Reconstruct the campaign's :class:`Experiment` and parameter-name order.

        Returns ``(experiment, parameter_names)``. Parameter names come
        from the :class:`ExperimentModel` so they match the order used
        by FIM / covariance matrices throughout this module.
        """
        from discopt.doe.templates import build_template

        meta = self.metadata()
        template = meta.get("template") or ""
        specs = self.input_specs()
        response = self.response_name()
        sigma = self.measurement_error()
        if template:
            args = self.template_args()
            degree = args.get("degree")
            inputs = [(s.name, s.lb, s.ub) for s in specs]
            exp = build_template(
                template,
                inputs=inputs,
                response_name=response,
                measurement_error=sigma,
                degree=int(degree) if degree is not None else None,
            )
        else:
            mod_call = self.module_callable()
            if not mod_call:
                raise ValueError(
                    "workbook has no template and no module_callable; cannot rebuild experiment"
                )
            exp = _load_module_callable(mod_call)

        em = exp.create_model(**self.param_initial_guess())
        return exp, em.parameter_names


def _load_module_callable(spec: str) -> Experiment:
    """Load ``"pkg.mod:callable"`` and return the callable's result.

    The callable must take no required arguments and return an
    :class:`Experiment` instance.
    """
    import importlib

    if ":" not in spec:
        raise ValueError(f"--module must be of the form 'pkg.mod:callable', got {spec!r}")
    mod_name, _, attr = spec.partition(":")
    if not mod_name or not attr:
        raise ValueError(f"invalid module spec {spec!r}")
    mod = importlib.import_module(mod_name)
    try:
        target = getattr(mod, attr)
    except AttributeError as e:
        raise ValueError(f"{mod_name!r} has no attribute {attr!r}") from e
    obj = target() if callable(target) else target
    if not isinstance(obj, Experiment):
        raise TypeError(
            f"{spec!r} did not produce an Experiment instance (got {type(obj).__name__})"
        )
    return obj


__all__ = [
    "InputSpec",
    "Workbook",
    "SHEET_ANOVA",
    "SHEET_FIM",
    "SHEET_HISTORY",
    "SHEET_INSTRUCTIONS",
    "SHEET_METADATA",
    "SHEET_PARAMETERS",
    "SHEET_RUNS",
]
