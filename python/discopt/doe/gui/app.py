"""Streamlit app for ``discopt doe``.

Launched by ``discopt doe gui [WORKBOOK]``. Reads the target workbook
path from ``$DISCOPT_DOE_WORKBOOK`` if set; otherwise the user picks
or creates one from the sidebar.

The app is a UI shell over the same pure ``do_*`` functions the CLI
uses. Every interaction reads campaign state through ``do_status`` so
buttons enable themselves only when their preconditions hold —
making the UI robust to partially-filled response columns, missing
fit results, or a workbook that was edited in Excel underneath us.
"""

from __future__ import annotations

import json
import os
from pathlib import Path
from typing import Any, Literal, cast

import openpyxl
import pandas as pd
import streamlit as st

from discopt.doe.cli import (
    DoEError,
    ExtendParams,
    NewParams,
    OptimizeParams,
    _design_row,
    do_extend,
    do_fit,
    do_new,
    do_optimize,
    do_status,
)
from discopt.doe.workbook import SHEET_ANOVA, SHEET_HISTORY, SHEET_RUNS, Workbook

_WORKBOOK_ENV = "DISCOPT_DOE_WORKBOOK"
_LOGO_PATH = Path(__file__).with_name("discopt-logo.png")
_ISSUES_URL = "https://github.com/jkitchin/discopt/issues"


# ──────────────────────────────────────────────────────────────────
# Helpers
# ──────────────────────────────────────────────────────────────────


def _init_state() -> None:
    if "workbook_path" not in st.session_state:
        env_path = os.environ.get(_WORKBOOK_ENV, "").strip()
        if env_path and Path(env_path).is_file():
            st.session_state["workbook_path"] = env_path
        else:
            st.session_state["workbook_path"] = None
    if "last_error" not in st.session_state:
        st.session_state["last_error"] = None


def _set_workbook(path: str | Path | None) -> None:
    st.session_state["workbook_path"] = str(path) if path else None
    st.session_state["last_error"] = None


def _safe_status(path: str) -> dict[str, Any] | None:
    try:
        return do_status({"workbook": path})
    except (FileNotFoundError, ValueError, DoEError) as e:
        st.session_state["last_error"] = str(e)
        return None


def _runs_df(path: str) -> pd.DataFrame:
    """Read the full runs sheet as a DataFrame for the editable table."""
    wb = openpyxl.load_workbook(path)
    sheet = wb[SHEET_RUNS]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return pd.DataFrame()
    headers = [str(c) for c in rows[0]]
    data = [r for r in rows[1:] if r and r[0] is not None]
    return pd.DataFrame(data, columns=headers)


_RESERVED_COLUMN_NAMES = frozenset({"run_id", "batch", "measured_at"})


def _rename_columns(
    path: str,
    *,
    input_renames: dict[str, str],
    response_rename: tuple[str, str] | None,
) -> None:
    """Rename input/response columns in an existing campaign.

    Updates the runs-sheet header row, the metadata ``input_specs`` /
    ``response_name`` / ``template_args.levels`` entries, and wipes the
    fit-derived sheets (``parameters``, ``fim``, ``anova``) since their
    row/column labels reference the old names. The user must re-run
    ``fit`` to repopulate them.
    """
    from discopt.doe.workbook import SHEET_FIM, SHEET_METADATA, SHEET_PARAMETERS

    wb = openpyxl.load_workbook(path)

    runs = wb[SHEET_RUNS]
    header_row = next(iter(runs.iter_rows(min_row=1, max_row=1)), ())
    for cell in header_row:
        val = cell.value
        if val in input_renames:
            cell.value = input_renames[val]
        elif response_rename and val == response_rename[0]:
            cell.value = response_rename[1]

    meta = wb[SHEET_METADATA]
    for row in meta.iter_rows(min_row=2):
        key = row[0].value
        val = row[1].value
        if key == "input_specs" and val:
            specs = json.loads(val)
            for s in specs:
                if s.get("name") in input_renames:
                    s["name"] = input_renames[s["name"]]
            row[1].value = json.dumps(specs, sort_keys=True)
        elif key == "response_name" and response_rename and val == response_rename[0]:
            row[1].value = response_rename[1]
        elif key == "template_args" and val:
            args = json.loads(val)
            levels = args.get("levels")
            if isinstance(levels, dict):
                args["levels"] = {input_renames.get(k, k): v for k, v in levels.items()}
                row[1].value = json.dumps(args, sort_keys=True)

    # Wipe fit artifacts. parameters keeps its header row; fim/anova are rewritten
    # from scratch on next fit so we clear them completely.
    params_sheet = wb[SHEET_PARAMETERS]
    if params_sheet.max_row > 1:
        params_sheet.delete_rows(2, params_sheet.max_row - 1)
    fim_sheet = wb[SHEET_FIM]
    if fim_sheet.max_row >= 1:
        fim_sheet.delete_rows(1, fim_sheet.max_row)
    if SHEET_ANOVA in wb.sheetnames:
        anova_sheet = wb[SHEET_ANOVA]
        if anova_sheet.max_row >= 1:
            anova_sheet.delete_rows(1, anova_sheet.max_row)

    wb.save(path)


def _save_responses(path: str, edited: pd.DataFrame, response: str) -> int:
    """Write edited response values back. Returns count of cells updated."""
    wb = openpyxl.load_workbook(path)
    sheet = wb[SHEET_RUNS]
    headers = [c.value for c in sheet[1]]
    resp_idx = headers.index(response)
    id_idx = headers.index("run_id")
    edits_by_id = {
        int(row["run_id"]): row[response] for _, row in edited.iterrows() if pd.notna(row["run_id"])
    }
    n_updates = 0
    for row in sheet.iter_rows(min_row=2):
        if row[id_idx].value is None:
            continue
        rid = int(row[id_idx].value)
        new_val = edits_by_id.get(rid)
        old_val = row[resp_idx].value
        if pd.isna(new_val) if new_val is not None else True:
            new_val = None
        if new_val != old_val:
            row[resp_idx].value = float(new_val) if new_val is not None else None
            n_updates += 1
    wb.save(path)
    return n_updates


def _model_terms(
    template: str | None,
    template_args: dict[str, Any],
    input_names: list[str],
    parameter_names: list[str],
) -> dict[str, str]:
    """Map each parameter name to its basis term (e.g. ``b12`` → ``x1·x2``).

    Returns an empty dict for templates we don't recognise (custom modules).
    The strings use a centered dot for products and Unicode superscripts
    for powers — they render cleanly in Streamlit tables and markdown.
    """
    if not template:
        return {}

    sup = str.maketrans("0123456789", "⁰¹²³⁴⁵⁶⁷⁸⁹")

    def _pow(name: str, exp: int) -> str:
        if exp == 0:
            return "1"
        if exp == 1:
            return name
        return f"{name}{str(exp).translate(sup)}"

    if template == "linear":
        out = {"b0": "1 (intercept)"}
        for i, nm in enumerate(input_names, start=1):
            out[f"b{i}"] = nm
        return out

    if template == "polynomial-1d":
        x = input_names[0] if input_names else "x"
        degree = int(template_args.get("degree", max(0, len(parameter_names) - 1)))
        return {f"b{j}": _pow(x, j) for j in range(degree + 1)}

    if template in ("response-surface-2d", "response-surface-3d"):
        n = len(input_names)
        out = {"b0": "1 (intercept)"}
        for i, nm in enumerate(input_names, start=1):
            out[f"b{i}"] = nm
        for i, nm in enumerate(input_names, start=1):
            out[f"b{i}{i}"] = _pow(nm, 2)
        for i in range(n):
            for j in range(i + 1, n):
                out[f"b{i + 1}{j + 1}"] = f"{input_names[i]}·{input_names[j]}"
        return out

    if template in ("scheffe-linear", "scheffe-quadratic", "scheffe-special-cubic"):
        q = len(input_names)
        out_s: dict[str, str] = {}
        for i, nm in enumerate(input_names, start=1):
            out_s[f"b{i}"] = nm
        if template == "scheffe-linear":
            return out_s
        for i in range(q):
            for j in range(i + 1, q):
                out_s[f"b{i + 1}{j + 1}"] = f"{input_names[i]}·{input_names[j]}"
        if template == "scheffe-quadratic":
            return out_s
        for i in range(q):
            for j in range(i + 1, q):
                for k in range(j + 1, q):
                    out_s[f"b{i + 1}{j + 1}{k + 1}"] = (
                        f"{input_names[i]}·{input_names[j]}·{input_names[k]}"
                    )
        return out_s

    return {}


def _model_equation(
    template: str | None,
    template_args: dict[str, Any],
    input_names: list[str],
    response: str,
    parameter_names: list[str],
    estimates: dict[str, float] | None = None,
) -> str | None:
    """Build a human-readable model equation.

    If ``estimates`` is supplied, numeric coefficients are substituted
    in place of symbolic names. Returns ``None`` for unrecognised
    templates.
    """
    terms = _model_terms(template, template_args, input_names, parameter_names)
    if not terms:
        return None

    def _coef(name: str, *, is_first: bool) -> str:
        if estimates is None or name not in estimates:
            return ("" if is_first else "+ ") + name
        v = float(estimates[name])
        sign = "" if is_first and v >= 0 else ("+ " if v >= 0 else "- ")
        return f"{sign}{abs(v):.4g}"

    parts: list[str] = []
    for i, name in enumerate(parameter_names):
        term = terms.get(name, name)
        coef = _coef(name, is_first=(i == 0))
        if term == "1 (intercept)" or term == "1":
            parts.append(coef.rstrip())
        else:
            parts.append(f"{coef}·{term}" if estimates is not None else f"{coef}·{term}")
    return f"{response} = " + " ".join(parts)


def _read_anova_sheet(path: str) -> dict[str, Any] | None:
    """Parse the 3-section ``anova`` sheet into structured pieces.

    Returns ``{"coefficients": DataFrame, "anova": DataFrame,
    "fit_summary": list[(label, value)]}`` or ``None`` if the sheet is
    missing or empty (e.g. workbook never fit).
    """
    wb = openpyxl.load_workbook(path)
    if SHEET_ANOVA not in wb.sheetnames:
        return None
    sheet = wb[SHEET_ANOVA]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return None
    # Sections are demarcated by single-cell section headers ("Coefficients",
    # "ANOVA", "Fit summary"). Walk the rows and split.
    sections: dict[str, list[tuple]] = {}
    current: str | None = None
    for row in rows:
        if row is None or all(c is None for c in row):
            continue
        first = row[0]
        rest = [c for c in row[1:] if c is not None]
        if (
            isinstance(first, str)
            and not rest
            and first in ("Coefficients", "ANOVA", "Fit summary")
        ):
            current = first
            sections[current] = []
            continue
        if current is not None:
            sections[current].append(row)

    def _to_df(rows_):
        if not rows_:
            return pd.DataFrame()
        header = [str(c) if c is not None else "" for c in rows_[0]]
        # Trim trailing empty columns
        while header and header[-1] == "":
            header.pop()
        body = [list(r[: len(header)]) for r in rows_[1:]]
        return pd.DataFrame(body, columns=header)

    fit_summary_rows = sections.get("Fit summary", [])
    fit_summary = [(str(r[0]), r[1]) for r in fit_summary_rows if r and r[0] is not None]

    return {
        "coefficients": _to_df(sections.get("Coefficients", [])),
        "anova": _to_df(sections.get("ANOVA", [])),
        "fit_summary": fit_summary,
    }


def _compute_parity(path: str, status: dict[str, Any]) -> pd.DataFrame | None:
    """Return per-completed-run observed/predicted/residual values.

    ``None`` if no fit results are available, no completed runs exist,
    or the experiment uses ``module_callable`` (no analytic design row).
    """
    template = status.get("template")
    if not template:
        return None
    params = status.get("parameters") or []
    if not params:
        return None
    try:
        wb_obj = Workbook.open(Path(path))
    except (FileNotFoundError, ValueError):
        return None
    completed = wb_obj.completed_runs()
    if not completed:
        return None
    parameter_names = [p["name"] for p in params]
    beta = [float(p["estimate"]) for p in params]
    input_names = [s.name for s in wb_obj.input_specs()]
    template_args = wb_obj.template_args()
    response = status["response_name"]

    rows = []
    for run in completed:
        x_row = _design_row(template, template_args, parameter_names, input_names, run)
        y_pred = float(sum(b * x for b, x in zip(beta, x_row)))
        y_obs = float(run[response])
        rows.append(
            {
                "run_id": int(run["run_id"]),
                "batch": int(run["batch"]) if run.get("batch") is not None else None,
                "y_observed": y_obs,
                "y_predicted": y_pred,
                "residual": y_obs - y_pred,
            }
        )
    return pd.DataFrame(rows)


def _history_rows(path: str) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(path)
    if SHEET_HISTORY not in wb.sheetnames:
        return []
    sheet = wb[SHEET_HISTORY]
    rows = list(sheet.iter_rows(values_only=True))
    if len(rows) < 2:
        return []
    headers = [str(c) for c in rows[0]]
    return [dict(zip(headers, r)) for r in rows[1:] if r and r[0] is not None]


# ──────────────────────────────────────────────────────────────────
# Sidebar
# ──────────────────────────────────────────────────────────────────


# GUI template menu collapses response-surface-2d and -3d into a single
# "response surface" choice; we pick the underlying template by factor
# count when generating.
_GUI_TEMPLATE_CHOICES = (
    "linear",
    "polynomial-1d",
    "response-surface",
    "mixture (Scheffé)",
    "Latin square (ANOVA)",
    "Factor screening (2-level)",
    "Active-learning optimization",
)
_GUI_TEMPLATE_HELP = {
    "linear": (
        "First-order model: y = b0 + b1·x1 + ... + bn·xn. "
        "Use for screening — does the response depend on each factor "
        "at all? Works with any number of factors."
    ),
    "polynomial-1d": (
        "Single-factor polynomial: y = b0 + b1·x + b2·x² + ... + bd·xᵈ. "
        "Use when you have one factor and need to capture curvature. "
        "Pick the degree below."
    ),
    "response-surface": (
        "Full-quadratic response surface in 2 or 3 factors: intercept, "
        "main effects, pure quadratic terms, and pairwise interactions. "
        "Use for optimization once you know the factors matter."
    ),
    "mixture (Scheffé)": (
        "Scheffé canonical mixture model: components are blend "
        "proportions / volumes constrained to sum to a fixed total. "
        "Pick the polynomial order (linear, quadratic, special cubic) "
        "and the total below."
    ),
    "Latin square (ANOVA)": (
        "Classical Latin / Graeco-Latin / hyper-Graeco-Latin square for "
        "ANOVA on additive main effects. 3 factors → Latin (k² runs), "
        "4 → Graeco-Latin, 5 → hyper-Graeco-Latin. Each factor has the "
        "same number of categorical levels."
    ),
    "Factor screening (2-level)": (
        "Answer 'does this factor matter?' with a full 2-level "
        "factorial. Each factor is varied between a LOW and HIGH "
        "level; 2^k runs at the corners. Add center points to detect "
        "curvature; replicate to estimate σ. Use this before fitting "
        "a response surface."
    ),
    "Active-learning optimization": (
        "Find the input that minimizes or maximizes the response using "
        "as few experiments as possible. Each round fits a surrogate "
        "(GP, response surface, or your own UQ-providing sklearn model) "
        "to completed runs and recommends the next batch by acquisition "
        "(expected improvement, UCB, steepest ascent)."
    ),
}

_LATIN_FACTOR_COUNT_TO_TEMPLATE = {
    3: "latin-square",
    4: "graeco-latin",
    5: "hyper-graeco-latin",
}

_SCHEFFE_FORMS = {
    "linear": "scheffe-linear",
    "quadratic": "scheffe-quadratic",
    "special-cubic": "scheffe-special-cubic",
}


def _sidebar() -> None:
    st.sidebar.title("discopt doe")

    mode_options = ["Open existing", "New campaign"]
    default_idx = 0 if st.session_state.get("workbook_path") else 1
    mode = st.sidebar.radio(
        "Workbook",
        mode_options,
        index=default_idx,
        help=(
            "**Open existing** loads an `.xlsx` campaign that's already "
            "on disk. **New campaign** generates a fresh optimal design "
            "and writes a new workbook."
        ),
    )

    if mode == "Open existing":
        _sidebar_open()
    else:
        _sidebar_new()


def _native_file_picker() -> str | None:
    """Spawn a native OS file picker as a subprocess and return the path.

    Returns ``None`` if the user cancels, no picker is available, or the
    subprocess fails. Cross-platform: AppleScript on macOS, ``zenity`` /
    ``kdialog`` on Linux, PowerShell ``OpenFileDialog`` on Windows.

    Filters by `*.xlsx` *visually*, but allows the user to switch to "all
    files" — the legacy 4-char OSType filter that grays out modern .xlsx
    files on macOS is *not* used.
    """
    import shutil
    import subprocess
    import sys

    if sys.platform == "darwin":
        script = (
            "try\n"
            "    set theFile to choose file with prompt "
            '"Open discopt doe workbook"\n'
            "    return POSIX path of theFile\n"
            "on error\n"
            '    return ""\n'
            "end try\n"
        )
        osascript = shutil.which("osascript")
        if not osascript:
            return None
        try:
            r = subprocess.run(
                [osascript, "-e", script],
                capture_output=True,
                text=True,
                timeout=3600,
            )
        except (subprocess.TimeoutExpired, FileNotFoundError):
            return None
        return r.stdout.strip() or None

    if sys.platform.startswith("linux"):
        zenity = shutil.which("zenity")
        if zenity:
            try:
                r = subprocess.run(
                    [
                        zenity,
                        "--file-selection",
                        "--title=Open discopt doe workbook",
                        "--file-filter=Excel workbook | *.xlsx",
                        "--file-filter=All files | *",
                    ],
                    capture_output=True,
                    text=True,
                    timeout=3600,
                )
            except (subprocess.TimeoutExpired, FileNotFoundError):
                return None
            return r.stdout.strip() or None
        kdialog = shutil.which("kdialog")
        if kdialog:
            try:
                r = subprocess.run(
                    [
                        kdialog,
                        "--getopenfilename",
                        ".",
                        "Excel workbook (*.xlsx)|All files (*)",
                    ],
                    capture_output=True,
                    text=True,
                    timeout=3600,
                )
            except (subprocess.TimeoutExpired, FileNotFoundError):
                return None
            return r.stdout.strip() or None
        return None

    if sys.platform == "win32":
        ps = (
            "Add-Type -AssemblyName System.Windows.Forms; "
            "$d = New-Object System.Windows.Forms.OpenFileDialog; "
            "$d.Title = 'Open discopt doe workbook'; "
            "$d.Filter = 'Excel workbook (*.xlsx)|*.xlsx|All files (*.*)|*.*'; "
            "if ($d.ShowDialog() -eq 'OK') { Write-Output $d.FileName }"
        )
        powershell = shutil.which("powershell") or shutil.which("pwsh")
        if not powershell:
            return None
        try:
            r = subprocess.run(
                [powershell, "-NoProfile", "-Command", ps],
                capture_output=True,
                text=True,
                timeout=3600,
            )
        except (subprocess.TimeoutExpired, FileNotFoundError):
            return None
        return r.stdout.strip() or None

    return None


def _sidebar_open() -> None:
    current = st.session_state.get("workbook_path") or ""
    path_input = st.sidebar.text_input(
        "Path to .xlsx",
        value=current,
        help=(
            "Absolute or `~`-relative path to a `discopt doe` workbook. "
            "The file must already exist."
        ),
    )
    col_open, col_browse = st.sidebar.columns(2)
    if col_open.button(
        "Open",
        help="Load the workbook at the path above into the GUI.",
        use_container_width=True,
    ):
        candidate = Path(path_input).expanduser().resolve()
        if not candidate.is_file():
            st.sidebar.error(f"Not found: {candidate}")
        else:
            _set_workbook(candidate)
            st.rerun()
    if col_browse.button(
        "Browse…",
        help="Open a native OS file picker.",
        use_container_width=True,
    ):
        picked = _native_file_picker()
        if picked is None:
            st.sidebar.caption("_(no file selected)_")
        else:
            candidate = Path(picked).expanduser().resolve()
            if candidate.is_file():
                _set_workbook(candidate)
                st.rerun()
            else:
                st.sidebar.error(f"Not found: {candidate}")


def _sidebar_new() -> None:
    template_choice = st.sidebar.selectbox(
        "Model template",
        _GUI_TEMPLATE_CHOICES,
        help=(
            "The model the design will be optimal for. Hover the option "
            "and read the caption below for what each template means."
        ),
    )
    st.sidebar.caption(_GUI_TEMPLATE_HELP[template_choice])

    scheffe_form: str | None = None
    mixture_total: float | None = None
    if template_choice == "Latin square (ANOVA)":
        _sidebar_new_latin()
        return
    if template_choice == "Factor screening (2-level)":
        _sidebar_new_factorial()
        return
    if template_choice == "Active-learning optimization":
        _sidebar_new_optimize()
        return
    if template_choice == "polynomial-1d":
        degree = st.sidebar.number_input(
            "Polynomial degree",
            min_value=1,
            max_value=10,
            value=2,
            help=(
                "Highest power of x in the model. Degree 1 = linear, "
                "2 = quadratic, 3 = cubic. The fit has `degree + 1` "
                "coefficients."
            ),
        )
        n_factors_default = 1
        factors_fixed = True
    elif template_choice == "response-surface":
        degree = None
        n_factors_default = 2
        factors_fixed = False
    elif template_choice == "mixture (Scheffé)":
        degree = None
        scheffe_form = st.sidebar.selectbox(
            "Mixture model order",
            list(_SCHEFFE_FORMS.keys()),
            index=1,
            help=(
                "Scheffé canonical polynomial. **linear** = pure-blend "
                "terms only (q parameters); **quadratic** adds pairwise "
                "blending (q + q(q-1)/2 parameters); **special-cubic** "
                "adds three-way blending (requires q ≥ 3)."
            ),
        )
        mixture_total = float(
            st.sidebar.number_input(
                "Sum to (mixture total)",
                min_value=1e-9,
                value=1.0,
                format="%g",
                help=(
                    "Required sum of the component values. Use 1.0 for "
                    "fractions, or an absolute total (mass, volume) when "
                    "factor bounds are in physical units."
                ),
            )
        )
        n_factors_default = 3
        factors_fixed = False
    else:  # linear
        degree = None
        n_factors_default = 2
        factors_fixed = False

    st.sidebar.markdown("**Factors**")
    if template_choice == "response-surface":
        st.sidebar.caption(
            "Add exactly 2 or 3 rows — one per design factor. Each needs "
            "a name and lower/upper bounds defining its experimental range."
        )
    elif template_choice == "linear":
        st.sidebar.caption(
            "One row per design factor. Add as many as you want; each needs "
            "a name and lower/upper bounds."
        )
    elif template_choice == "mixture (Scheffé)":
        st.sidebar.caption(
            "One row per mixture component (at least 2; at least 3 for "
            "special-cubic). Bounds default to [0, sum-to]; tighten them "
            "for realistic per-component limits. The sum equality is "
            "enforced automatically."
        )
    else:
        st.sidebar.caption(
            "One row for the single design factor: name plus the range "
            "the polynomial will be fit over."
        )
    if template_choice == "mixture (Scheffé)":
        ub_default = mixture_total if mixture_total is not None else 1.0
        default_factors = pd.DataFrame(
            {
                "name": ["A", "B", "C"][:n_factors_default]
                + [f"x{i + 1}" for i in range(3, n_factors_default)],
                "lb": [0.0] * n_factors_default,
                "ub": [float(ub_default)] * n_factors_default,
            }
        )
    else:
        default_factors = pd.DataFrame(
            {
                "name": [f"x{i + 1}" for i in range(n_factors_default)],
                "lb": [0.0] * n_factors_default,
                "ub": [1.0] * n_factors_default,
            }
        )
    num_rows: Literal["fixed", "dynamic"] = "fixed" if factors_fixed else "dynamic"
    factors_df = st.sidebar.data_editor(
        default_factors,
        num_rows=num_rows,
        key=f"factors_editor_{template_choice}",
        width="stretch",
        column_config={
            "name": st.column_config.TextColumn(
                "name", help="Factor name; becomes a column header in the workbook."
            ),
            "lb": st.column_config.NumberColumn(
                "lb", help="Lower bound of the factor's experimental range."
            ),
            "ub": st.column_config.NumberColumn("ub", help="Upper bound (must exceed lb)."),
        },
    )

    response_name = st.sidebar.text_input(
        "Response name",
        value="y",
        help=(
            "Name of the measured quantity (e.g. 'yield', 'conversion'). "
            "Becomes the response column header in the workbook."
        ),
    )
    n = st.sidebar.number_input(
        "Initial runs",
        min_value=1,
        value=6,
        help=(
            "How many experiments to generate in this first batch. Need "
            "at least as many runs as parameters for a well-posed fit."
        ),
    )
    error = st.sidebar.number_input(
        "Measurement noise σ",
        min_value=1e-9,
        value=1.0,
        format="%g",
        help=(
            "Estimated standard deviation of the measurement error in the "
            "response, in the same units as the response. Used to weight "
            "the Fisher Information Matrix."
        ),
    )
    criterion = st.sidebar.selectbox(
        "Optimality criterion",
        ["determinant", "trace", "min_eigenvalue", "condition_number"],
        help=(
            "**determinant** (D-optimal, default) maximizes the FIM's "
            "log-determinant — minimizes confidence-ellipsoid volume. "
            "**trace** (A) minimizes average parameter variance. "
            "**min_eigenvalue** (E) maximizes worst-case parameter info. "
            "**condition_number** balances information across directions."
        ),
    )
    seed = st.sidebar.number_input(
        "Random seed",
        value=42,
        step=1,
        help="Seed for the multi-start optimizer; change for a different design.",
    )
    n_starts = st.sidebar.number_input(
        "Multi-start budget",
        min_value=1,
        value=5,
        help=(
            "Number of random initializations for each design point's "
            "optimization. Higher = more thorough search, slower."
        ),
    )

    output_dir, output_name = _output_path_picker()
    output_path = str(Path(output_dir) / output_name)

    if st.sidebar.button(
        "Generate initial design",
        type="primary",
        help=(
            "Solve the optimal-design problem, write the workbook to disk, and open it in this GUI."
        ),
    ):
        inputs = _validate_factors(factors_df)
        if inputs is None:
            return
        actual_template = _resolve_template(template_choice, len(inputs), scheffe_form)
        if actual_template is None:
            if template_choice == "mixture (Scheffé)":
                st.sidebar.error(
                    "Scheffé mixture needs at least 2 components "
                    "(at least 3 for special-cubic); got "
                    f"{len(inputs)}."
                )
            else:
                st.sidebar.error(
                    f"Response surface needs exactly 2 or 3 factors; got {len(inputs)}."
                )
            return
        out_path = Path(output_path).expanduser().resolve()
        if out_path.exists():
            st.sidebar.error(f"{out_path} already exists. Pick another path.")
            return
        params = NewParams(
            output=out_path,
            n=int(n),
            inputs=inputs,
            response_name=response_name,
            measurement_error=float(error),
            criterion=criterion,
            seed=int(seed),
            n_starts=int(n_starts),
            template=actual_template,
            degree=int(degree) if degree is not None else None,
            mixture_total=mixture_total,
        )
        try:
            with st.spinner("Solving optimal design..."):
                do_new(params)
        except (DoEError, ValueError) as e:
            st.sidebar.error(str(e))
            return
        _set_workbook(out_path)
        st.rerun()


def _sidebar_new_latin() -> None:
    """Sidebar widgets for creating a Latin-family workbook."""
    n_factors = int(
        st.sidebar.selectbox(
            "Number of factors",
            (3, 4, 5),
            index=0,
            help=(
                "3 = Latin square (1 treatment + 2 blocking factors). "
                "4 = Graeco-Latin square. 5 = hyper-Graeco-Latin square."
            ),
        )
    )
    template = _LATIN_FACTOR_COUNT_TO_TEMPLATE[n_factors]
    st.sidebar.caption(f"Generates a `{template}` design.")

    k = int(
        st.sidebar.number_input(
            "Levels per factor (k)",
            min_value=2,
            max_value=9,
            value=4,
            help=(
                "Number of levels each factor takes. The design uses k² "
                "runs per replicate. For 4+ factors k = 6 is rejected "
                "(no MOLS pair exists)."
            ),
        )
    )
    replicates = int(
        st.sidebar.number_input(
            "Replicates",
            min_value=1,
            max_value=20,
            value=1,
            help=(
                "Number of independent randomizations of the whole "
                "square. Replication is needed when k is small to give "
                "the residual MS enough degrees of freedom for ANOVA."
            ),
        )
    )

    st.sidebar.markdown("**Factors and levels**")
    st.sidebar.caption(
        "One row per factor. Each `levels` cell is a comma-separated "
        "list of k values (numeric or string)."
    )
    default_names = ["row", "col", "treatment", "block_D", "block_E"][:n_factors]
    default_levels = ",".join(str(i + 1) for i in range(k))
    default_df = pd.DataFrame(
        {
            "name": default_names,
            "levels": [default_levels] * n_factors,
        }
    )
    factors_df = st.sidebar.data_editor(
        default_df,
        num_rows="fixed",
        key=f"latin_factors_editor_{n_factors}_{k}",
        width="stretch",
        column_config={
            "name": st.column_config.TextColumn(
                "name", help="Factor name; becomes a column header."
            ),
            "levels": st.column_config.TextColumn(
                "levels",
                help="Comma-separated level list. All factors must have exactly k levels.",
            ),
        },
    )

    response_name = st.sidebar.text_input("Response name", value="y")
    seed = st.sidebar.number_input("Random seed", value=42, step=1)

    output_dir, output_name = _output_path_picker()
    output_path = str(Path(output_dir) / output_name)

    if st.sidebar.button(
        "Generate Latin design",
        type="primary",
        help="Build the Latin square, write the workbook, and open it here.",
    ):
        levels: dict[str, list[object]] = {}
        for _, row in factors_df.iterrows():
            name = str(row["name"]).strip()
            if not name:
                st.sidebar.error("Every factor must have a name.")
                return
            raw = [p.strip() for p in str(row["levels"]).split(",") if p.strip()]
            if len(raw) != k:
                st.sidebar.error(f"Factor {name!r}: expected {k} levels, got {len(raw)}.")
                return
            try:
                vals: list[object] = [float(p) for p in raw]
            except ValueError:
                vals = list(raw)
            levels[name] = vals
        if len(levels) != n_factors:
            st.sidebar.error(f"Need {n_factors} unique factor names.")
            return
        if n_factors >= 4 and k == 6:
            st.sidebar.error("k = 6 is not allowed for 4+ factor Latin designs.")
            return

        out_path = Path(output_path).expanduser().resolve()
        if out_path.exists():
            st.sidebar.error(f"{out_path} already exists. Pick another path.")
            return

        params = NewParams(
            output=out_path,
            n=k * k * replicates,
            inputs=[],
            response_name=response_name,
            measurement_error=1.0,
            criterion="anova",
            seed=int(seed),
            n_starts=1,
            template=template,
            levels=levels,
            replicates=replicates,
        )
        try:
            with st.spinner("Building Latin design..."):
                do_new(params)
        except (DoEError, ValueError) as e:
            st.sidebar.error(str(e))
            return
        _set_workbook(out_path)
        st.rerun()


def _sidebar_new_factorial() -> None:
    """Sidebar widgets for creating a 2-level factorial workbook."""
    st.sidebar.markdown("**Factors (LOW / HIGH)**")
    st.sidebar.caption(
        "One row per factor (2 - 8 rows). LOW / HIGH may be numbers "
        "(e.g. 80, 120) or strings (e.g. A, B). Each factor's two "
        "values define its two test levels."
    )
    default_df = pd.DataFrame(
        {
            "name": ["temp", "pressure", "catalyst"],
            "low": ["80", "1", "A"],
            "high": ["120", "5", "B"],
        }
    )
    factors_df = st.sidebar.data_editor(
        default_df,
        num_rows="dynamic",
        key="factorial_factors_editor",
        width="stretch",
        column_config={
            "name": st.column_config.TextColumn(
                "name", help="Factor name; becomes a column header."
            ),
            "low": st.column_config.TextColumn("low", help="LOW level (numeric or string)."),
            "high": st.column_config.TextColumn("high", help="HIGH level (numeric or string)."),
        },
    )

    center_points = int(
        st.sidebar.number_input(
            "Center points (per replicate)",
            min_value=0,
            max_value=20,
            value=0,
            help=(
                "Number of center-point runs added per replicate. "
                "Requires all factors to be numeric. Useful to detect "
                "curvature and estimate σ without assuming the linear "
                "model is correct."
            ),
        )
    )
    replicates = int(
        st.sidebar.number_input(
            "Replicates",
            min_value=1,
            max_value=10,
            value=1,
            help=(
                "Number of independent replications of the whole "
                "design (each replicate is freshly randomized)."
            ),
        )
    )

    response_name = st.sidebar.text_input("Response name", value="y")
    seed = st.sidebar.number_input("Random seed", value=42, step=1)

    output_dir, output_name = _output_path_picker()
    output_path = str(Path(output_dir) / output_name)

    if st.sidebar.button(
        "Generate screening design",
        type="primary",
        help="Build the 2-level factorial, write the workbook, open it here.",
    ):

        def _coerce(v: str) -> object:
            try:
                return float(v)
            except (TypeError, ValueError):
                return v

        pairs: dict[str, tuple[object, object]] = {}
        for _, row in factors_df.iterrows():
            name = str(row["name"]).strip()
            if not name:
                continue
            lo_raw = str(row["low"]).strip()
            hi_raw = str(row["high"]).strip()
            if not lo_raw or not hi_raw:
                st.sidebar.error(f"Factor {name!r}: both LOW and HIGH required.")
                return
            lo, hi = _coerce(lo_raw), _coerce(hi_raw)
            if lo == hi:
                st.sidebar.error(f"Factor {name!r}: LOW and HIGH must differ.")
                return
            pairs[name] = (lo, hi)
        if len(pairs) < 2 or len(pairs) > 8:
            st.sidebar.error(f"Need 2 - 8 factors, got {len(pairs)}.")
            return
        if center_points > 0:
            non_numeric = [
                n
                for n, (lo, hi) in pairs.items()
                if not (
                    isinstance(lo, (int, float))
                    and isinstance(hi, (int, float))
                    and not isinstance(lo, bool)
                )
            ]
            if non_numeric:
                st.sidebar.error(
                    f"Center points require numeric factors only; non-numeric: "
                    f"{', '.join(non_numeric)}."
                )
                return

        out_path = Path(output_path).expanduser().resolve()
        if out_path.exists():
            st.sidebar.error(f"{out_path} already exists. Pick another path.")
            return
        params = NewParams(
            output=out_path,
            n=2 ** len(pairs) * replicates + center_points * replicates,
            inputs=[],
            response_name=response_name,
            measurement_error=1.0,
            criterion="anova",
            seed=int(seed),
            n_starts=1,
            template="factorial-2level",
            factor_pairs=pairs,
            center_points=center_points,
            replicates=replicates,
        )
        try:
            with st.spinner("Building factorial design..."):
                do_new(params)
        except (DoEError, ValueError) as e:
            st.sidebar.error(str(e))
            return
        _set_workbook(out_path)
        st.rerun()


def _sidebar_new_optimize() -> None:
    """Sidebar widgets for creating an active-learning optimization workbook."""
    st.sidebar.markdown("**Factors (box bounds)**")
    st.sidebar.caption(
        "One row per numeric factor. Each row needs a name and the "
        "lower / upper bounds defining the search box. Categorical "
        "factors are not yet supported here — encode them as 0/1 "
        "indicators upstream."
    )
    default_factors = pd.DataFrame(
        {
            "name": ["x1", "x2"],
            "lb": [-1.0, -1.0],
            "ub": [1.0, 1.0],
        }
    )
    factors_df = st.sidebar.data_editor(
        default_factors,
        num_rows="dynamic",
        key="optimize_factors_editor",
        width="stretch",
        column_config={
            "name": st.column_config.TextColumn("name"),
            "lb": st.column_config.NumberColumn("lb"),
            "ub": st.column_config.NumberColumn("ub"),
        },
    )

    response_name = st.sidebar.text_input("Response name", value="y")
    measurement_error = float(
        st.sidebar.number_input(
            "Measurement noise σ",
            min_value=1e-9,
            value=1.0,
            format="%g",
            help=(
                "Estimated standard deviation of the response. Stored "
                "in the workbook for reference; the surrogate handles "
                "noise on its own."
            ),
        )
    )

    st.sidebar.markdown("**Initial design**")
    n_seed = int(
        st.sidebar.number_input(
            "Seed batch size",
            min_value=2,
            value=8,
            help=(
                "Sobol-distributed points inside the box, run once "
                "before active learning takes over. A good default is "
                "2--5 × the number of factors."
            ),
        )
    )

    st.sidebar.markdown("**Active-learning defaults**")
    criterion = st.sidebar.selectbox(
        "Objective",
        ["maximize", "minimize"],
        help="Are we looking for the highest or lowest response?",
    )
    surrogate = st.sidebar.selectbox(
        "Surrogate (default)",
        ["gp", "response-surface"],
        help=(
            "**gp**: Gaussian process with Matern(5/2) + white noise. "
            "Best for smooth nonlinear responses with < 200 runs. "
            "**response-surface**: degree-2 polynomial + Bayesian "
            "ridge. Faster, interpretable, restricted to quadratic."
        ),
    )
    acquisition = st.sidebar.selectbox(
        "Acquisition (default)",
        ["expected_improvement", "ucb", "steepest_ascent"],
        help=(
            "**expected_improvement** balances mean and uncertainty "
            "(the standard BO choice). **ucb** uses μ + κσ; tune κ in "
            "the main panel. **steepest_ascent** ignores σ — useful "
            "with a response-surface surrogate for classical "
            "Box-Wilson behavior."
        ),
    )

    seed = st.sidebar.number_input("Random seed", value=42, step=1)
    output_dir, output_name = _output_path_picker()
    output_path = str(Path(output_dir) / output_name)

    if st.sidebar.button(
        "Generate seed batch",
        type="primary",
        help="Sobol-sample the seed batch, write the workbook, open it.",
    ):
        inputs = _validate_factors(factors_df)
        if inputs is None:
            return
        if len(inputs) < 1:
            st.sidebar.error("Need at least one factor.")
            return
        out_path = Path(output_path).expanduser().resolve()
        if out_path.exists():
            st.sidebar.error(f"{out_path} already exists. Pick another path.")
            return
        params = NewParams(
            output=out_path,
            n=int(n_seed),
            inputs=inputs,
            response_name=response_name,
            measurement_error=measurement_error,
            criterion="active-learning",
            seed=int(seed),
            n_starts=1,
            template="optimize",
            optimize_criterion=criterion,
            optimize_surrogate=surrogate,
            optimize_acquisition=acquisition,
        )
        try:
            with st.spinner("Sobol-sampling the seed batch..."):
                do_new(params)
        except (DoEError, ValueError) as e:
            st.sidebar.error(str(e))
            return
        _set_workbook(out_path)
        st.rerun()


def _resolve_template(
    gui_choice: str, n_factors: int, scheffe_form: str | None = None
) -> str | None:
    """Map the GUI's template menu choice to an actual `do_new` template."""
    if gui_choice == "response-surface":
        if n_factors == 2:
            return "response-surface-2d"
        if n_factors == 3:
            return "response-surface-3d"
        return None
    if gui_choice == "mixture (Scheffé)":
        if scheffe_form is None or scheffe_form not in _SCHEFFE_FORMS:
            return None
        if n_factors < 2:
            return None
        if scheffe_form == "special-cubic" and n_factors < 3:
            return None
        return _SCHEFFE_FORMS[scheffe_form]
    return gui_choice


def _output_path_picker() -> tuple[str, str]:
    """Sidebar widget: directory browser + filename → (dir, filename)."""
    st.sidebar.markdown("**Output workbook**")

    if "output_dir" not in st.session_state:
        st.session_state["output_dir"] = str(Path.cwd())

    current_dir = Path(st.session_state["output_dir"]).expanduser()
    if not current_dir.is_dir():
        current_dir = Path.cwd()
        st.session_state["output_dir"] = str(current_dir)

    dir_input = st.sidebar.text_input(
        "Folder",
        value=str(current_dir),
        key="output_dir_input",
        help=(
            "Folder where the new `.xlsx` will be written. Type a path "
            "directly, or use **Browse folder…** below to navigate. "
            "`~` is expanded."
        ),
    )
    typed = Path(dir_input).expanduser()
    if typed.is_dir() and str(typed.resolve()) != str(current_dir.resolve()):
        st.session_state["output_dir"] = str(typed.resolve())
        current_dir = typed.resolve()

    with st.sidebar.expander("Browse folder…", expanded=False):
        st.caption(f"`{current_dir}`")
        try:
            entries = sorted(current_dir.iterdir(), key=lambda p: p.name.lower())
            subdirs = [p for p in entries if p.is_dir() and not p.name.startswith(".")]
            xlsx_files = [p for p in entries if p.is_file() and p.suffix.lower() == ".xlsx"]
        except (OSError, PermissionError) as e:
            st.error(f"Cannot list folder: {e}")
            subdirs, xlsx_files = [], []

        if st.button(
            "⬆ Parent folder",
            key="browse_parent",
            disabled=current_dir.parent == current_dir,
            help="Navigate up one directory.",
        ):
            st.session_state["output_dir"] = str(current_dir.parent)
            st.rerun()

        if subdirs:
            st.caption("Subfolders")
            for d in subdirs[:50]:
                if st.button(f"📁 {d.name}", key=f"browse_dir_{d}"):
                    st.session_state["output_dir"] = str(d)
                    st.rerun()
            if len(subdirs) > 50:
                st.caption(f"…and {len(subdirs) - 50} more (type the path above).")
        else:
            st.caption("_No subfolders._")

        if xlsx_files:
            st.caption("Existing `.xlsx` files (read-only preview):")
            for f in xlsx_files[:20]:
                st.text(f"• {f.name}")

    output_name = st.sidebar.text_input(
        "Filename",
        value="campaign.xlsx",
        key="output_name_input",
        help=(
            "Name of the new workbook (must end in `.xlsx`). Must not "
            "already exist in the chosen folder."
        ),
    )

    full = Path(st.session_state["output_dir"]) / output_name
    st.sidebar.caption(f"→ `{full}`")
    return st.session_state["output_dir"], output_name


def _validate_factors(df: pd.DataFrame) -> list[tuple[str, float, float]] | None:
    rows: list[tuple[str, float, float]] = []
    for i, row in df.iterrows():
        name = str(row.get("name") or "").strip()
        lb, ub = row.get("lb"), row.get("ub")
        all_blank = not name and pd.isna(lb) and pd.isna(ub)
        if all_blank:
            continue
        if not name:
            st.sidebar.error(f"Row {int(i) + 1}: factor name is required.")
            return None
        if pd.isna(lb) or pd.isna(ub):
            st.sidebar.error(f"Factor {name!r}: both lower and upper bounds are required.")
            return None
        if float(ub) <= float(lb):
            st.sidebar.error(f"Factor {name!r}: upper bound must exceed lower bound.")
            return None
        rows.append((name, float(lb), float(ub)))
    if not rows:
        st.sidebar.error("Define at least one factor.")
        return None
    return rows


# ──────────────────────────────────────────────────────────────────
# Main pane
# ──────────────────────────────────────────────────────────────────


def _main_pane() -> None:
    path = st.session_state.get("workbook_path")
    if not path:
        _empty_state()
        return

    status = _safe_status(path)
    if status is None:
        st.error(st.session_state.get("last_error") or "Could not open workbook.")
        if st.button("Forget this workbook"):
            _set_workbook(None)
            st.rerun()
        return

    _status_banner(status)
    _rename_panel(path, status)
    _runs_editor(path, status)
    template = status.get("template") or ""
    if template in {"latin-square", "graeco-latin", "hyper-graeco-latin", "factorial-2level"}:
        _anova_panel(path, status)
    elif template == "optimize":
        _optimize_panel(path, status)
    else:
        _fit_panel(path, status)
        _extend_panel(path, status)
    _history_panel(path)
    _download_panel(path)


def _empty_state() -> None:
    st.title("discopt doe")
    st.markdown(
        "Pick **Open existing** to load an `.xlsx` campaign, or **New "
        "campaign** to generate a fresh optimal design."
    )
    st.markdown(
        "The CLI shortcut is `discopt doe gui PATH.xlsx` — that drops you "
        "straight into the right workbook."
    )


def _status_banner(status: dict[str, Any]) -> None:
    label = status["template"] or status["module_callable"] or "(unknown model)"
    st.subheader(f"Workbook: `{Path(status['workbook_path']).name}`")
    cols = st.columns(4)
    cols[0].metric("Model", label)
    cols[1].metric("Completed", f"{status['n_completed']} / {status['n_total']}")
    cols[2].metric("Pending", status["n_pending"])
    cols[3].metric(
        "Parameters",
        len(status["parameters"]) if status["parameters"] else 0,
    )

    template = status.get("template") or ""
    template_args = status.get("template_args") or {}
    if template in {
        "latin-square",
        "graeco-latin",
        "hyper-graeco-latin",
        "factorial-2level",
    }:
        levels_map = template_args.get("levels", {})
        inputs_str = ", ".join(
            f"{name} ∈ {{{', '.join(str(v) for v in vals)}}}" for name, vals in levels_map.items()
        )
    else:
        inputs_str = ", ".join(
            f"{s['name']} ∈ [{s['lb']}, {s['ub']}]" for s in status["input_specs"]
        )
    st.caption(f"**Response**: `{status['response_name']}`  •  **Factors**: {inputs_str}")

    input_names = [s["name"] for s in status["input_specs"]]
    parameter_names = [p["name"] for p in (status.get("parameters") or [])]
    equation = _model_equation(
        status.get("template"),
        status.get("template_args") or {},
        input_names,
        status["response_name"],
        parameter_names,
    )
    if equation:
        st.caption(f"**Model**: `{equation}`")
    st.caption(f"**Next step (CLI)**: `{status['next_command']}`")


def _rename_panel(path: str, status: dict[str, Any]) -> None:
    """Expander UI to rename factor / response columns of an open campaign.

    Per option B in the design discussion: rename succeeds even if a fit
    already exists, but the fit artifacts (parameters / FIM / ANOVA) are
    cleared and the user must re-run ``fit``.
    """
    input_specs = status.get("input_specs") or []
    old_response = status["response_name"]
    old_inputs = [s["name"] for s in input_specs]
    has_fit = bool(status.get("parameters")) or bool(status.get("n_parameters"))

    with st.expander("Rename factors / response", expanded=False):
        st.caption(
            "Rename the response or any factor column. Updates the `runs` "
            "header, `metadata`, and template levels. `run_id`, `batch`, "
            "and `measured_at` are reserved and cannot be renamed."
        )
        if has_fit:
            st.warning(
                "This workbook has fit results. Renaming will clear the "
                "`parameters`, `fim`, and `anova` sheets — you'll need to "
                "re-run **Fit** afterwards."
            )

        new_response = st.text_input(
            "Response",
            value=old_response,
            key=f"rename_response_{path}",
        )
        new_inputs: list[str] = []
        for i, old in enumerate(old_inputs):
            new_inputs.append(
                st.text_input(
                    f"Factor {i + 1}",
                    value=old,
                    key=f"rename_factor_{i}_{path}",
                )
            )

        if not st.button("Apply renames", key=f"rename_apply_{path}"):
            return

        # Validate.
        candidates = [new_response, *new_inputs]
        for c in candidates:
            if not c or not c.strip():
                st.error("Names cannot be empty.")
                return
        stripped = [c.strip() for c in candidates]
        if len(set(stripped)) != len(stripped):
            st.error("New names must be unique (response + every factor).")
            return
        for c in stripped:
            if c in _RESERVED_COLUMN_NAMES:
                st.error(f"`{c}` is reserved (run_id / batch / measured_at).")
                return

        new_response_s = stripped[0]
        new_inputs_s = stripped[1:]
        input_renames = {old: new for old, new in zip(old_inputs, new_inputs_s) if old != new}
        response_rename: tuple[str, str] | None = (
            (old_response, new_response_s) if old_response != new_response_s else None
        )
        if not input_renames and response_rename is None:
            st.info("No name changes detected.")
            return

        try:
            _rename_columns(
                path,
                input_renames=input_renames,
                response_rename=response_rename,
            )
        except (FileNotFoundError, ValueError, KeyError, OSError) as e:
            st.error(f"Rename failed: {e}")
            return

        msg_parts = []
        if input_renames:
            msg_parts.append("factors: " + ", ".join(f"{a}→{b}" for a, b in input_renames.items()))
        if response_rename:
            msg_parts.append(f"response: {response_rename[0]}→{response_rename[1]}")
        st.success("Renamed " + " · ".join(msg_parts) + ".")
        if has_fit:
            st.info("Fit artifacts cleared — run **Fit** again to repopulate.")
        st.rerun()


def _runs_editor(path: str, status: dict[str, Any]) -> None:
    st.markdown("### Runs")
    df = _runs_df(path)
    if df.empty:
        st.info("No runs yet. Generate an initial design from the sidebar.")
        return

    response = status["response_name"]
    input_cols = [s["name"] for s in status["input_specs"]]
    locked = [c for c in df.columns if c not in (response,)]
    input_col_configs: dict[str, Any] = {}
    for c in input_cols:
        col_dtype = df[c].dtype if c in df.columns else None
        if col_dtype is not None and pd.api.types.is_numeric_dtype(col_dtype):
            input_col_configs[c] = st.column_config.NumberColumn(c, disabled=True)
        else:
            input_col_configs[c] = st.column_config.TextColumn(c, disabled=True)
    edited = st.data_editor(
        df,
        disabled=locked,
        num_rows="fixed",
        width="stretch",
        key=f"runs_editor_{path}",
        column_config={
            response: st.column_config.NumberColumn(
                response, help=f"Measured value for '{response}' (blank = pending)"
            ),
            **input_col_configs,
        },
    )

    col_a, col_b = st.columns([1, 1])
    if col_a.button("Save responses", type="primary"):
        n = _save_responses(path, edited, response)
        if n == 0:
            st.info("No changes detected.")
        else:
            st.success(f"Updated {n} cell(s).")
            st.rerun()
    if col_b.button("Reload from disk"):
        st.rerun()


def _anova_panel(path: str, status: dict[str, Any]) -> None:
    """ANOVA panel shown for Latin-family workbooks in place of fit/extend."""
    from discopt.doe.cli import do_anova

    st.markdown("### ANOVA")
    n_done = status["n_completed"]
    n_total = status["n_total"]
    if n_done == 0:
        st.info(f"No completed runs yet — fill in the response column (0/{n_total}).")
        return
    if n_done < n_total:
        st.warning(
            f"{n_done} of {n_total} runs have responses. ANOVA needs all "
            "runs (or at least enough for residual df > 0)."
        )

    levels_map = (status.get("template_args") or {}).get("levels") or {}
    factor_names = list(levels_map.keys())
    interactions: list[tuple[str, ...]] = []
    if len(factor_names) >= 2:
        choices = [":".join(p) for p in _two_way_pairs(factor_names)]
        if choices:
            picked = st.multiselect(
                "Include 2-way interactions",
                choices,
                default=[],
                help=(
                    "Optional pairwise interaction terms. Each consumes "
                    "(k-1)·(k-1) residual degrees of freedom; available "
                    "only if there are enough df left."
                ),
            )
            interactions = [tuple(s.split(":")) for s in picked]
    include_replicate = False
    replicates = int((status.get("template_args") or {}).get("replicates", 1) or 1)
    if replicates > 1:
        include_replicate = st.checkbox(
            "Treat replicate as a blocking factor",
            value=False,
            help="Adds the replicate index as an additional ANOVA factor.",
        )

    if st.button(
        f"Run ANOVA ({n_done} observation(s))",
        type="primary",
        key="anova_btn",
    ):
        out: dict[str, Any] | None
        try:
            out = do_anova(
                {
                    "workbook": path,
                    "interactions": interactions or None,
                    "include_replicate": include_replicate,
                }
            )
        except (DoEError, ValueError) as e:
            st.error(str(e))
            return
        st.session_state["last_anova_result"] = out

    out = cast("dict[str, Any] | None", st.session_state.get("last_anova_result"))
    if not out:
        return
    cols = st.columns(3)
    cols[0].metric("Observations", out["n_observations"])
    cols[1].metric("Grand mean", f"{out['grand_mean']:.4g}")
    cols[2].metric("Balanced", "yes" if out["balanced"] else "no")
    rows_df = pd.DataFrame(out["rows"])
    rows_df = rows_df.rename(columns={"ss": "SS", "df": "df", "ms": "MS", "f": "F", "p": "p-value"})
    st.dataframe(rows_df, width="stretch")
    if not out["balanced"]:
        st.caption("Design is unbalanced; Type-I SS depend on factor order in the workbook.")


def _two_way_pairs(names: list[str]) -> list[tuple[str, str]]:
    out = []
    for i, a in enumerate(names):
        for b in names[i + 1 :]:
            out.append((a, b))
    return out


def _optimize_panel(path: str, status: dict[str, Any]) -> None:
    """Active-learning round driver shown for optimize-template workbooks."""
    st.markdown("### Active-learning round")

    template_args = status.get("template_args") or {}
    default_criterion = template_args.get("criterion", "maximize")
    default_surrogate = template_args.get("surrogate", "gp")
    default_acquisition = template_args.get("acquisition", "expected_improvement")
    input_names = [s["name"] for s in status["input_specs"]]

    if status["n_completed"] == 0:
        st.info(
            "No completed runs yet — fill in the response column above "
            "and save before requesting the first active-learning batch."
        )
        return

    col1, col2 = st.columns(2)
    criterion = col1.selectbox(
        "Objective",
        ["maximize", "minimize"],
        index=0 if default_criterion == "maximize" else 1,
        help="Are we looking for the highest or lowest response?",
    )
    batch_size = int(
        col2.number_input(
            "Batch size",
            min_value=1,
            max_value=64,
            value=4,
            help="Number of new experiments to recommend this round.",
        )
    )

    col3, col4 = st.columns(2)
    surrogate_choice = col3.selectbox(
        "Surrogate",
        ["gp", "response-surface", "custom (sklearn import)"],
        index=(
            0
            if default_surrogate == "gp"
            else (1 if default_surrogate == "response-surface" else 2)
        ),
        help=(
            "**gp**: Matern(5/2) + white noise. **response-surface**: "
            "degree-2 polynomial + Bayesian ridge. **custom**: import "
            "any UQ-providing sklearn-style estimator below."
        ),
    )
    acquisition = col4.selectbox(
        "Acquisition",
        ["expected_improvement", "ucb", "steepest_ascent"],
        index={"expected_improvement": 0, "ucb": 1, "steepest_ascent": 2}.get(
            default_acquisition, 0
        ),
        help=(
            "**expected_improvement** balances mean and uncertainty. "
            "**ucb** uses μ + κσ (tune κ below). "
            "**steepest_ascent** ignores σ — pair with response-surface "
            "for Box-Wilson behavior."
        ),
    )

    custom_path: str | None = None
    custom_kwargs: dict[str, Any] | None = None
    if surrogate_choice == "custom (sklearn import)":
        st.markdown("**Custom surrogate**")
        st.caption(
            "Importable dotted path to an estimator class. The adapter "
            "auto-detects `predict(X, return_std=True)`, "
            "`predict(X, return_interval=True)`, or falls back to a "
            "residual bootstrap. Constructor kwargs are JSON."
        )
        custom_path = st.text_input(
            "Estimator path",
            value="sklearn.gaussian_process:GaussianProcessRegressor",
            help="e.g. `pycse.sklearn.lpr:LinearLPR` or `sklearn.linear_model:BayesianRidge`",
        )
        kwargs_text = st.text_area(
            "Constructor kwargs (JSON)",
            value="{}",
            height=80,
            help="JSON dict forwarded to the class constructor.",
        )
        try:
            custom_kwargs = json.loads(kwargs_text) if kwargs_text.strip() else {}
            if not isinstance(custom_kwargs, dict):
                raise ValueError("must be a JSON object")
        except (json.JSONDecodeError, ValueError) as e:
            st.error(f"kwargs JSON is invalid: {e}")
            custom_kwargs = None

    acq_kwargs: dict[str, float] = {}
    if acquisition in {"ucb", "lcb", "confidence_bound"}:
        acq_kwargs["kappa"] = float(
            st.number_input(
                "UCB κ (explore/exploit tradeoff)",
                min_value=0.0,
                value=2.0,
                step=0.1,
                help="Larger κ = more exploration.",
            )
        )
    if acquisition in {"expected_improvement", "ei"}:
        acq_kwargs["xi"] = float(
            st.number_input(
                "EI ξ (min improvement)",
                min_value=0.0,
                value=0.0,
                step=0.001,
                format="%.4f",
                help=(
                    "Demand at least this much improvement before EI "
                    "rates a candidate. ξ > 0 encourages exploration."
                ),
            )
        )

    seed = int(
        st.number_input(
            "Random seed",
            value=int(status.get("seed", 0)) + 1,
            step=1,
            help="Seed for the Sobol candidate pool.",
        )
    )

    can_run = surrogate_choice != "custom (sklearn import)" or (
        custom_path and custom_kwargs is not None
    )
    if st.button(
        "Recommend next batch",
        type="primary",
        disabled=not can_run,
        help=(
            "Fit the surrogate to completed runs, score Sobol "
            "candidates, append the top batch to the workbook."
        ),
    ):
        params = OptimizeParams(
            workbook=Path(path),
            criterion=criterion,
            surrogate="gp" if surrogate_choice != "custom (sklearn import)" else default_surrogate,
            acquisition=acquisition,
            batch_size=batch_size,
            seed=seed,
            acquisition_kwargs=acq_kwargs or None,
            custom_surrogate_path=custom_path,
            custom_surrogate_kwargs=custom_kwargs,
        )
        if surrogate_choice == "response-surface":
            params.surrogate = "response-surface"
        out: dict[str, Any] | None
        try:
            with st.spinner("Fitting surrogate + scoring candidates..."):
                out = do_optimize(params)
        except (DoEError, ValueError, TypeError) as e:
            st.error(str(e))
            return
        st.session_state["last_optimize_result"] = out
        st.success(f"Appended {len(out['next_designs'])} new pending runs.")
        st.rerun()

    out = cast("dict[str, Any] | None", st.session_state.get("last_optimize_result"))
    if out and out.get("workbook_path") == str(Path(path)):
        st.markdown("**Last recommendation**")
        cols = st.columns(3)
        inc_y = out.get("incumbent_y")
        cols[0].metric(
            "Incumbent y",
            f"{inc_y:.4g}" if inc_y is not None else "—",
        )
        cols[1].metric("Surrogate mode", out.get("surrogate_mode") or "—")
        cols[2].metric("New runs", len(out.get("next_designs", [])))

        inc_x = out.get("incumbent_x")
        if inc_x:
            st.caption("Incumbent x: " + ", ".join(f"{k}={v:.4g}" for k, v in inc_x.items()))
        next_df = pd.DataFrame(out.get("next_designs", []))
        if not next_df.empty:
            st.dataframe(next_df, width="stretch", hide_index=True)

    _optimize_history_plot(path, status, input_names)


def _optimize_history_plot(path: str, status: dict[str, Any], input_names: list[str]) -> None:
    """Best-so-far curve + (for 1D/2D) a surrogate posterior plot."""
    from discopt.doe.workbook import Workbook as _Wb

    completed = _Wb.open(Path(path)).completed_runs()
    if len(completed) < 2:
        return
    response = status["response_name"]
    try:
        ys = [float(r[response]) for r in completed]
    except (KeyError, TypeError, ValueError):
        return
    template_args = status.get("template_args") or {}
    criterion = template_args.get("criterion", "maximize")
    direction = 1 if criterion == "maximize" else -1

    best_so_far = []
    cur = ys[0]
    for y in ys:
        cur = y if direction * y > direction * cur else cur
        best_so_far.append(cur)

    import matplotlib.pyplot as plt

    fig, ax = plt.subplots(figsize=(6, 3))
    ax.plot(
        range(1, len(ys) + 1), best_so_far, "o-", color="C0", label=f"best so far ({criterion})"
    )
    ax.scatter(range(1, len(ys) + 1), ys, color="grey", alpha=0.5, s=20, label="each run")
    ax.set_xlabel("run number")
    ax.set_ylabel(response)
    ax.set_title("Convergence so far")
    ax.legend()
    st.pyplot(fig, clear_figure=True)

    if len(input_names) == 1 and len(completed) >= 3:
        _plot_surrogate_1d(completed, input_names[0], response, status)
    elif len(input_names) == 2 and len(completed) >= 4:
        _plot_surrogate_2d(completed, input_names, response, status)


def _plot_surrogate_1d(
    completed: list[dict[str, Any]],
    xname: str,
    response: str,
    status: dict[str, Any],
) -> None:
    import matplotlib.pyplot as plt
    import numpy as _np

    from discopt.doe.surrogate import coerce_surrogate

    X = _np.array([[float(r[xname])] for r in completed])
    y = _np.array([float(r[response]) for r in completed])
    try:
        s = coerce_surrogate("gp")
        s.fit(X, y)
    except Exception as e:  # pragma: no cover - GP convergence noise
        st.caption(f"(could not fit surrogate for plot: {e})")
        return
    spec = next(s for s in status["input_specs"] if s["name"] == xname)
    xs = _np.linspace(spec["lb"], spec["ub"], 200).reshape(-1, 1)
    mu, sd = s.predict(xs)
    fig, ax = plt.subplots(figsize=(6, 3))
    ax.plot(xs.ravel(), mu, color="C0", label="GP mean")
    ax.fill_between(
        xs.ravel(), mu - 1.96 * sd, mu + 1.96 * sd, color="C0", alpha=0.2, label="95% band"
    )
    ax.scatter(X.ravel(), y, color="k", marker="o", label="runs")
    ax.set_xlabel(xname)
    ax.set_ylabel(response)
    ax.set_title("Surrogate (GP) posterior")
    ax.legend()
    st.pyplot(fig, clear_figure=True)


def _plot_surrogate_2d(
    completed: list[dict[str, Any]],
    names: list[str],
    response: str,
    status: dict[str, Any],
) -> None:
    import matplotlib.pyplot as plt
    import numpy as _np

    from discopt.doe.surrogate import coerce_surrogate

    X = _np.array([[float(r[names[0]]), float(r[names[1]])] for r in completed])
    y = _np.array([float(r[response]) for r in completed])
    try:
        s = coerce_surrogate("gp")
        s.fit(X, y)
    except Exception as e:  # pragma: no cover
        st.caption(f"(could not fit surrogate for plot: {e})")
        return
    spec_a = next(s for s in status["input_specs"] if s["name"] == names[0])
    spec_b = next(s for s in status["input_specs"] if s["name"] == names[1])
    g1 = _np.linspace(spec_a["lb"], spec_a["ub"], 60)
    g2 = _np.linspace(spec_b["lb"], spec_b["ub"], 60)
    G1, G2 = _np.meshgrid(g1, g2)
    Xg = _np.column_stack([G1.ravel(), G2.ravel()])
    mu, _ = s.predict(Xg)
    Mu = mu.reshape(G1.shape)
    fig, ax = plt.subplots(figsize=(6, 4.5))
    cs = ax.contourf(G1, G2, Mu, levels=20, cmap="viridis")
    plt.colorbar(cs, ax=ax, label=f"GP mean ({response})")
    ax.scatter(X[:, 0], X[:, 1], color="white", edgecolor="k", s=40, label="runs")
    ax.set_xlabel(names[0])
    ax.set_ylabel(names[1])
    ax.set_title("Surrogate (GP) posterior mean")
    ax.legend()
    st.pyplot(fig, clear_figure=True)


def _fit_panel(path: str, status: dict[str, Any]) -> None:
    st.markdown("### Fit")
    can_fit = status["n_completed"] > 0 and status["template"] is not None

    if not can_fit:
        if status["module_callable"]:
            st.info(
                "Fitting for `--module` experiments is not yet supported from "
                "the GUI. Use `discopt.estimate.estimate_parameters` in Python."
            )
        else:
            st.info("No completed runs yet — fill in some response values first.")
        return

    if st.button(f"Fit ({status['n_completed']} observation(s))", type="primary", key="fit_btn"):
        try:
            with st.spinner("Fitting parameters..."):
                do_fit({"workbook": path})
        except (DoEError, ValueError) as e:
            st.error(str(e))
            return
        st.rerun()

    if not status["parameters"]:
        return

    anova = _read_anova_sheet(path)
    parity = _compute_parity(path, status)
    _render_fit_results(status, anova, parity)


def _render_fit_results(
    status: dict[str, Any],
    anova: dict[str, Any] | None,
    parity: pd.DataFrame | None,
) -> None:
    """Render the post-fit diagnostics: coefficients, ANOVA, parity, residuals."""
    tabs = st.tabs(["Coefficients", "ANOVA", "Parity", "Residuals"])

    # --- Coefficients ---
    input_names = [s["name"] for s in status["input_specs"]]
    parameter_names = [p["name"] for p in status["parameters"]]
    estimates = {p["name"]: p["estimate"] for p in status["parameters"]}
    term_map = _model_terms(
        status.get("template"),
        status.get("template_args") or {},
        input_names,
        parameter_names,
    )

    with tabs[0]:
        symbolic = _model_equation(
            status.get("template"),
            status.get("template_args") or {},
            input_names,
            status["response_name"],
            parameter_names,
        )
        fitted_eq = _model_equation(
            status.get("template"),
            status.get("template_args") or {},
            input_names,
            status["response_name"],
            parameter_names,
            estimates=estimates,
        )
        if symbolic:
            st.markdown(f"**Model form:** `{symbolic}`")
        if fitted_eq:
            st.markdown(f"**Fitted equation:** `{fitted_eq}`")

        coef_df = (anova or {}).get("coefficients") if anova else None
        if coef_df is None or coef_df.empty:
            # Fall back to the parameters sheet (no t-stat / p-value).
            coef_df = pd.DataFrame(
                [
                    {
                        "name": p["name"],
                        "estimate": p["estimate"],
                        "std_error": p["std_error"],
                        "ci_lower_95": p["ci_lower_95"],
                        "ci_upper_95": p["ci_upper_95"],
                    }
                    for p in status["parameters"]
                ]
            )

        # Insert a "term" column right after "name" so each coefficient
        # is annotated with the basis function it multiplies.
        if "name" in coef_df.columns and term_map:
            coef_df = coef_df.copy()
            coef_df.insert(1, "term", coef_df["name"].map(lambda n: term_map.get(str(n), "")))
        st.dataframe(coef_df, width="stretch", hide_index=True)
        st.caption(
            "`term` shows the basis function each coefficient multiplies. "
            "p-values test H₀: coefficient = 0 (two-sided t-test, "
            "df = n_obs − n_params). Lower p = stronger evidence the term "
            "is needed."
        )

    # --- ANOVA + fit summary ---
    with tabs[1]:
        if not anova or anova["anova"].empty:
            st.info("No ANOVA table yet — re-run `fit` to populate it.")
        else:
            st.markdown("**Decomposition**")
            st.dataframe(anova["anova"], width="stretch", hide_index=True)
            st.markdown("**Fit summary**")
            summary = anova["fit_summary"]
            if summary:
                # Promote the most useful stats to metric tiles.
                lookup = dict(summary)
                cols = st.columns(4)
                _metric(cols[0], "R²", lookup.get("R_squared"), fmt="{:.4f}")
                _metric(cols[1], "Adj. R²", lookup.get("adjusted_R_squared"), fmt="{:.4f}")
                _metric(cols[2], "RMSE", lookup.get("RMSE (sigma_hat)"), fmt="{:.4g}")
                _metric(
                    cols[3],
                    "F p-value",
                    lookup.get("F_p_value"),
                    fmt="{:.3g}",
                )
                st.dataframe(
                    pd.DataFrame(summary, columns=["statistic", "value"]),
                    width="stretch",
                    hide_index=True,
                )

    # --- Parity ---
    with tabs[2]:
        if parity is None or parity.empty:
            st.info("Parity plot needs completed runs and a fitted model.")
        else:
            st.altair_chart(_parity_chart(parity), use_container_width=True)
            st.caption(
                "Each point is one completed run; the diagonal is perfect "
                "prediction. Points far off the line indicate fit error or "
                "outliers."
            )

    # --- Residuals ---
    with tabs[3]:
        if parity is None or parity.empty:
            st.info("Residual plot needs completed runs and a fitted model.")
        else:
            st.altair_chart(_residual_chart(parity), use_container_width=True)
            st.caption(
                "Residual = observed − predicted, plotted vs predicted. "
                "Look for trends or fanning out (signs of model "
                "mis-specification or heteroscedasticity)."
            )


def _metric(col: Any, label: str, value: Any, *, fmt: str = "{}") -> None:
    if isinstance(value, (int, float)):
        try:
            col.metric(label, fmt.format(value))
            return
        except (ValueError, TypeError):
            pass
    col.metric(label, "—")


def _parity_chart(df: pd.DataFrame) -> Any:
    import altair as alt

    lo = float(min(df["y_observed"].min(), df["y_predicted"].min()))
    hi = float(max(df["y_observed"].max(), df["y_predicted"].max()))
    pad = 0.05 * (hi - lo or 1.0)
    domain = [lo - pad, hi + pad]

    points = (
        alt.Chart(df)
        .mark_circle(size=80, opacity=0.75)
        .encode(
            x=alt.X("y_predicted:Q", scale=alt.Scale(domain=domain), title="Predicted"),
            y=alt.Y("y_observed:Q", scale=alt.Scale(domain=domain), title="Observed"),
            color=alt.Color("batch:N", title="Batch"),
            tooltip=["run_id", "batch", "y_observed", "y_predicted", "residual"],
        )
    )
    diag = (
        alt.Chart(pd.DataFrame({"x": domain, "y": domain}))
        .mark_line(strokeDash=[4, 4], color="gray")
        .encode(x="x:Q", y="y:Q")
    )
    return (diag + points).properties(height=380)


def _residual_chart(df: pd.DataFrame) -> Any:
    import altair as alt

    points = (
        alt.Chart(df)
        .mark_circle(size=80, opacity=0.75)
        .encode(
            x=alt.X("y_predicted:Q", title="Predicted"),
            y=alt.Y("residual:Q", title="Residual (observed − predicted)"),
            color=alt.Color("batch:N", title="Batch"),
            tooltip=["run_id", "batch", "y_observed", "y_predicted", "residual"],
        )
    )
    zero = (
        alt.Chart(pd.DataFrame({"y": [0]}))
        .mark_rule(color="gray", strokeDash=[4, 4])
        .encode(y="y:Q")
    )
    return (zero + points).properties(height=380)


def _extend_panel(path: str, status: dict[str, Any]) -> None:
    st.markdown("### Extend (next D-optimal batch)")
    if status["n_pending"] > 0:
        st.info(
            f"{status['n_pending']} run(s) still pending. Fill them in and refit "
            "before extending so the new batch can build on what you learned."
        )
        return
    n = st.number_input("Number of new runs", min_value=1, value=4, key="extend_n")
    n_starts = st.number_input("Multi-start budget", min_value=1, value=5, key="extend_starts")
    if st.button(f"Append {int(n)} run(s)", type="primary", key="extend_btn"):
        try:
            with st.spinner("Solving next-batch design..."):
                do_extend(ExtendParams(workbook=Path(path), n=int(n), n_starts=int(n_starts)))
        except (DoEError, ValueError) as e:
            st.error(str(e))
            return
        st.rerun()


def _history_panel(path: str) -> None:
    rows = _history_rows(path)
    if not rows:
        return
    with st.expander(f"History ({len(rows)} entr{'y' if len(rows) == 1 else 'ies'})"):
        recent = rows[-10:]
        formatted = []
        for r in recent:
            args = r.get("args")
            if isinstance(args, str):
                try:
                    args = json.dumps(json.loads(args))
                except (ValueError, TypeError):
                    pass
            formatted.append(
                {
                    "timestamp": r.get("timestamp"),
                    "command": r.get("command"),
                    "args": args,
                }
            )
        st.dataframe(pd.DataFrame(formatted), width="stretch", hide_index=True)


def _download_panel(path: str) -> None:
    p = Path(path)
    if not p.is_file():
        return
    st.download_button(
        "Download workbook",
        data=p.read_bytes(),
        file_name=p.name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ──────────────────────────────────────────────────────────────────
# Entry point
# ──────────────────────────────────────────────────────────────────


def main() -> None:
    st.set_page_config(
        page_title="discopt doe",
        page_icon=str(_LOGO_PATH) if _LOGO_PATH.is_file() else None,
        layout="wide",
    )
    if _LOGO_PATH.is_file():
        st.logo(str(_LOGO_PATH), size="large", link=_ISSUES_URL.rsplit("/", 1)[0])
    _init_state()
    if _LOGO_PATH.is_file():
        st.sidebar.image(str(_LOGO_PATH), width=120)
    _sidebar()
    st.sidebar.divider()
    st.sidebar.markdown(
        f"[Report an issue]({_ISSUES_URL}) &nbsp;·&nbsp; "
        f"[discopt on GitHub]({_ISSUES_URL.rsplit('/', 1)[0]})"
    )
    _main_pane()


if __name__ == "__main__":
    main()
